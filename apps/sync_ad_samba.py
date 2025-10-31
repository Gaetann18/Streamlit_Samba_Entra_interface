def run():
    import streamlit as st
    import subprocess
    import random
    from contextlib import contextmanager
    import paramiko
    import sys
    import os
    sys.path.append('/home/streamlit')
    import config
    import logging
    from typing import Tuple, Dict, List
    import time
    import pandas as pd
    import json
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill
    import uuid
    import requests
    from msal import ConfidentialClientApplication
    # ========================= Configuration depuis config.py =========================
    # Récupération des configurations depuis le fichier config.py
    keytab_path = config.KEYTAB_PATH
    krb5_ccache = config.KRB5_CCACHE
    krb5_conf = config.KRB5_CONF
    venv_path = config.VENV_PATH
    sync_script = config.SYNC_SCRIPT
    config_file = config.AZURE_CONFIG_FILE
    CSV_FILE = config.CSV_FILE
    PASSWORD_EXCEL_FILE = config.PASSWORD_EXCEL_FILE

    # Variables d'environnement pour Kerberos et venv
    os.environ["KRB5CCNAME"] = krb5_ccache
    os.environ["KRB5_CONFIG"] = krb5_conf
    os.environ["KRB5_TRACE"] = "/dev/stderr"

    # Utilisation de la fonction utilitaire de config.py
    env = config.get_kerberos_env()

    # Logging
    logging.basicConfig(level=getattr(logging, config.LOG_LEVEL, logging.INFO), 
                       format=config.LOG_FORMAT)
    logger = logging.getLogger(__name__)
    
    # Mapping des classes depuis config.py
    CLASS_MAPPING = config.CLASS_MAPPING
    
    # ========================= Fonctions utilitaires =========================
    
    def get_samba_users_list():
        """Récupère la liste réelle des utilisateurs depuis Samba"""
        users_list = []
        
        try:
            with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                if client:
                    # Commande pour lister tous les utilisateurs samba
                    cmd = "sudo -S samba-tool user list"
                    output, error = execute_ssh_command(client, cmd, config.SAMBA_PWD)
                    
                    if output and not error:
                        # Filtrer les utilisateurs (enlever les comptes système)
                        ignored_users = config.get_ignored_users()
                        
                        for line in output.split('\n'):
                            username = line.strip()
                            if username and username.lower() not in [u.lower() for u in ignored_users]:
                                users_list.append(username)
                        
                        st.success(f"✅ {len(users_list)} utilisateurs trouvés dans Samba")
                    else:
                        st.error(f"❌ Erreur lors de la récupération des utilisateurs: {error}")
                        
        except Exception as e:
            st.error(f"❌ Erreur de connexion Samba: {e}")
            
        return users_list
    
    def get_samba_group_members(group_name):
        """Récupère la liste des membres d'un groupe Samba spécifique"""
        members_list = []
        
        try:
            with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                if client:
                    # Commande pour lister les membres d'un groupe
                    cmd = f"sudo -S samba-tool group listmembers {group_name}"
                    output, error = execute_ssh_command(client, cmd, config.SAMBA_PWD)
                    
                    if output and not error:
                        # Filtrer les utilisateurs
                        ignored_users = config.get_ignored_users()
                        
                        for line in output.split('\n'):
                            username = line.strip()
                            if username and username.lower() not in [u.lower() for u in ignored_users]:
                                members_list.append(username)
                        
                        st.info(f"📋 {len(members_list)} membres trouvés dans le groupe '{group_name}'")
                    else:
                        if "does not exist" in error.lower():
                            st.warning(f"⚠️ Le groupe '{group_name}' n'existe pas")
                        else:
                            st.error(f"❌ Erreur lors de la récupération du groupe: {error}")
                        
        except Exception as e:
            st.error(f"❌ Erreur de connexion Samba: {e}")
            
        return members_list

    def add_students_to_wifi_group_by_description(descriptions_filter, target_group="WIFI"):
        """Ajoute les élèves au groupe WIFI basé sur leurs descriptions (classe)"""
        results = []
        added_count = 0
        
        try:
            # Récupérer la liste des utilisateurs existants
            df_users = get_existing_users()
            
            if df_users.empty:
                st.warning("Aucun utilisateur trouvé dans la base de données")
                return results, 0
                
            # Filtrer les utilisateurs selon leurs descriptions (classes)
            matching_users = []
            for _, user in df_users.iterrows():
                if 'Classe' in df_users.columns:
                    user_classe = str(user['Classe']).strip().upper()
                    # Vérifier si la classe contient une des descriptions recherchées
                    for desc in descriptions_filter:
                        if desc.upper() in user_classe:
                            matching_users.append(user)
                            break
            
            if not matching_users:
                st.info(f"Aucun utilisateur trouvé avec les descriptions : {', '.join(descriptions_filter)}")
                return results, 0
            
            st.info(f"🔍 {len(matching_users)} utilisateurs trouvés correspondant aux critères")
            
            # Connexion SSH et ajout au groupe
            with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                if client:
                    # Créer le groupe WIFI s'il n'existe pas
                    create_group_cmd = f"sudo -S samba-tool group add {target_group} || true"
                    output, error = execute_ssh_command(client, create_group_cmd, config.SAMBA_PWD)
                    
                    # Ajouter chaque utilisateur au groupe WIFI
                    for user in matching_users:
                        username = user['Login']
                        user_class = user['Classe'] if 'Classe' in user else 'N/A'
                        
                        add_to_group_cmd = f"sudo -S samba-tool group addmembers {target_group} {username}"
                        output, error = execute_ssh_command(client, add_to_group_cmd, config.SAMBA_PWD)
                        
                        if error:
                            if "is already a member" in error.lower():
                                results.append(f"ℹ️ {username} ({user_class}) : Déjà membre du groupe {target_group}")
                            else:
                                results.append(f"❌ {username} ({user_class}) : {error}")
                        else:
                            results.append(f"✅ {username} ({user_class}) : Ajouté au groupe {target_group}")
                            added_count += 1
                else:
                    st.error("❌ Impossible de se connecter au serveur Samba")
                    
        except Exception as e:
            st.error(f"❌ Erreur lors de l'ajout au groupe WIFI : {e}")
            results.append(f"❌ Erreur générale : {e}")
            
        return results, added_count

    def get_imfr_students(config_dict=None):
        """Récupère la liste des élèves depuis IMFR en utilisant selenium_scraper"""
        try:
            # Import de la classe EleveScraper
            from selenium_scraper import EleveScraper
            
            # Utiliser la configuration fournie ou celle par défaut
            if config_dict is None:
                if hasattr(config, 'IMFR_CONFIG'):
                    config_dict = config.IMFR_CONFIG
                else:
                    raise Exception("Configuration IMFR non trouvée. Veuillez fournir les paramètres de connexion.")
            
            # Initialiser le scraper
            scraper = EleveScraper(headless=True)
            
            # Effectuer le scraping complet
            eleves_data = scraper.scrape_eleves_complete(config_dict)
            
            # Fermer le driver
            if scraper.driver:
                scraper.driver.quit()
            
            # Convertir en DataFrame
            if eleves_data:
                df_eleves = pd.DataFrame(eleves_data)
                
                # Normaliser les colonnes pour correspondre au format attendu
                if 'Nom_Complet' in df_eleves.columns and 'Nom' not in df_eleves.columns:
                    # Essayer de séparer nom et prénom depuis Nom_Complet
                    def split_nom_prenom(nom_complet):
                        parts = str(nom_complet).strip().split()
                        if len(parts) >= 2:
                            nom = parts[0].upper()
                            prenom = " ".join(parts[1:]).title()
                        else:
                            nom = str(nom_complet).upper()
                            prenom = ""
                        return pd.Series([nom, prenom])
                    
                    df_eleves[['Nom', 'Prénom']] = df_eleves['Nom_Complet'].apply(split_nom_prenom)
                
                # Assurer que les colonnes nécessaires existent
                required_columns = ['Nom', 'Prénom', 'Classe']
                for col in required_columns:
                    if col not in df_eleves.columns:
                        df_eleves[col] = 'Non défini'
                
                # Nettoyer les données
                df_eleves['Nom'] = df_eleves['Nom'].astype(str).str.strip().str.upper()
                df_eleves['Prénom'] = df_eleves['Prénom'].astype(str).str.strip().str.title()
                df_eleves['Classe'] = df_eleves['Classe'].astype(str).str.strip()
                
                # Supprimer les lignes vides
                df_eleves = df_eleves[
                    (df_eleves['Nom'] != '') & 
                    (df_eleves['Nom'] != 'NAN') & 
                    (df_eleves['Nom'].notna())
                ].copy()
                
                logger.info(f"IMFR: {len(df_eleves)} élèves récupérés avec succès")
                return df_eleves
            else:
                logger.warning("IMFR: Aucune donnée récupérée")
                return pd.DataFrame(columns=['Nom', 'Prénom', 'Classe'])
                
        except ImportError:
            raise Exception("Module selenium_scraper non trouvé. Veuillez vérifier l'installation.")
        except Exception as e:
            logger.error(f"Erreur lors de la récupération IMFR: {e}")
            raise Exception(f"Erreur lors de la récupération des données IMFR: {str(e)}")
    
    # ========================= Fonctions Microsoft Graph pour licences =========================
    
    @st.cache_data(ttl=60)  # Cache réduit à 1 minute pour debug
    def get_available_licenses() -> Dict[str, Dict]:
        """Récupère la liste des licences disponibles depuis Microsoft Graph"""
        licenses = {}
        
        try:
            app = ConfidentialClientApplication(
                config.CLIENT_ID,
                authority=f"https://login.microsoftonline.com/{config.TENANT_ID}",
                client_credential=config.CLIENT_SECRET
            )
            
            token_response = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
            access_token = token_response.get("access_token")
            
            if not access_token:
                st.error("Impossible d'obtenir le token Microsoft pour les licences")
                return {}

            headers = {'Authorization': f'Bearer {access_token}'}
            
            # Récupérer les SKUs de licences disponibles
            url = "https://graph.microsoft.com/v1.0/subscribedSkus"
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            skus_data = response.json().get("value", [])
            
            for sku in skus_data:
                sku_id = sku.get("skuId", "")
                sku_part_number = sku.get("skuPartNumber", "")
                display_name = sku.get("skuPartNumber", "")  # On utilise skuPartNumber comme nom d'affichage
                
                # Mapper les noms techniques vers des noms plus lisibles
                friendly_names = {
                    "STANDARDWOFFPACK_FACULTY": "Office 365 A1 for Faculty",
                    "STANDARDWOFFPACK_STUDENT": "Office 365 A1 for Students",
                    "ENTERPRISEPACK_FACULTY": "Office 365 A3 for Faculty",
                    "ENTERPRISEPACK_STUDENT": "Office 365 A3 for Students",
                    "ENTERPRISEPREMIUM_FACULTY": "Office 365 A5 for Faculty",
                    "ENTERPRISEPREMIUM_STUDENT": "Office 365 A5 for Students",
                    "OFFICESUBSCRIPTION": "Microsoft 365 Apps for Business",
                    "O365_BUSINESS_ESSENTIALS": "Microsoft 365 Business Basic",
                    "O365_BUSINESS_PREMIUM": "Microsoft 365 Business Premium",
                    "ENTERPRISEPACK": "Microsoft 365 E3",
                    "ENTERPRISEPREMIUM": "Microsoft 365 E5"
                }
                
                friendly_name = friendly_names.get(sku_part_number, display_name)
                
                # Obtenir les unités disponibles
                enabled_units = sku.get("prepaidUnits", {}).get("enabled", 0)
                consumed_units = sku.get("consumedUnits", 0)
                available_units = enabled_units - consumed_units
                
                if enabled_units > 0:  # Seulement les licences avec des unités disponibles
                    licenses[sku_id] = {
                        "id": sku_id,
                        "partNumber": sku_part_number,
                        "displayName": friendly_name,
                        "enabledUnits": enabled_units,
                        "consumedUnits": consumed_units,
                        "availableUnits": available_units
                    }
            
            logger.info(f"Récupéré {len(licenses)} types de licences depuis Microsoft Graph")
            
        except requests.RequestException as e:
            st.error(f"Erreur lors de la récupération des licences : {e}")
            logger.error(f"Erreur Microsoft Graph licences: {e}")
        except Exception as e:
            st.error(f"Erreur générale lors de la récupération des licences : {e}")
            logger.error(f"Erreur générale licences: {e}")
        
        return licenses
    
    @st.cache_data(ttl=300)  # Cache 5 minutes
    def get_entra_groups() -> Dict[str, Dict]:
        """Récupère la liste des groupes depuis Entra ID via Microsoft Graph"""
        groups = {}
        
        try:
            app = ConfidentialClientApplication(
                config.CLIENT_ID,
                authority=f"https://login.microsoftonline.com/{config.TENANT_ID}",
                client_credential=config.CLIENT_SECRET
            )
            
            token_response = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
            access_token = token_response.get("access_token")
            
            if not access_token:
                st.error("Impossible d'obtenir le token Microsoft pour les groupes")
                return {}

            headers = {'Authorization': f'Bearer {access_token}'}
            
            # Récupérer les groupes depuis Entra ID
            url = "https://graph.microsoft.com/v1.0/groups?$select=id,displayName,description,groupTypes,membershipRule,securityEnabled,mailEnabled&$top=999"
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            groups_data = response.json().get("value", [])
            
            for group in groups_data:
                # Validation et sécurisation des données
                if not group or not isinstance(group, dict):
                    continue
                    
                group_id = group.get("id")
                if not group_id:
                    continue
                    
                display_name = group.get("displayName") or "Sans nom"
                description = group.get("description")  # Peut être None
                group_types = group.get("groupTypes") or []
                security_enabled = bool(group.get("securityEnabled", False))
                mail_enabled = bool(group.get("mailEnabled", False))
                
                # Déterminer le type de groupe
                if "Unified" in group_types:
                    group_type = "Microsoft 365"
                elif security_enabled and not mail_enabled:
                    group_type = "Sécurité"
                elif mail_enabled and not security_enabled:
                    group_type = "Distribution"
                elif security_enabled and mail_enabled:
                    group_type = "Sécurité avec messagerie"
                else:
                    group_type = "Autre"
                
                # Sécuriser la description (peut être None)
                safe_description = description or ""
                truncated_description = safe_description[:100] + "..." if len(safe_description) > 100 else safe_description
                
                groups[group_id] = {
                    "id": group_id,
                    "displayName": display_name,
                    "description": truncated_description,
                    "groupType": group_type,
                    "securityEnabled": security_enabled,
                    "mailEnabled": mail_enabled
                }
            
            logger.info(f"Récupéré {len(groups)} groupes depuis Entra ID")
            
        except requests.RequestException as e:
            st.error(f"Erreur lors de la récupération des groupes : {e}")
            logger.error(f"Erreur Microsoft Graph groupes: {e}")
        except Exception as e:
            st.error(f"Erreur générale lors de la récupération des groupes : {e}")
            logger.error(f"Erreur générale groupes: {e}")
        
        return groups
    
    def normalize_class_name(original_class):
        """Normalise le nom de classe selon le mapping défini"""
        cleaned_class = " ".join(original_class.split())
        
        # Vérification directe dans le mapping
        if cleaned_class in CLASS_MAPPING:
            return CLASS_MAPPING[cleaned_class]
        
        # Traitement spécial pour les classes TFP
        if "TFP" in cleaned_class.upper():
            # Nettoyer et simplifier
            tfp_clean = cleaned_class.upper().replace("  ", " ").strip()
            
            # Identifier le type (TEAVA ou RAV)
            if "TEAVA" in tfp_clean:
                return "TFPTEAVA"
            elif "RAV" in tfp_clean or "Rav" in cleaned_class:
                return "TFPRAV"
            else:
                # TFP générique
                return "TFP"
        
        # Normalisation standard pour les autres classes
        normalized = cleaned_class.replace(" ", "").replace("è", "e").replace("nd", "").upper()
        logger.warning(f"Classe non mappée: '{original_class}' -> utilisation par défaut: '{normalized}'")
        return normalized
    
    def detect_csv_encoding(file_obj):
        """Détecte l'encodage d'un fichier CSV et retourne le DataFrame"""
        # Liste des encodages à tester, dans l'ordre de préférence
        encodings = config.CSV_ENCODINGS
        
        for encoding in encodings:
            try:
                # Remettre le pointeur au début du fichier
                file_obj.seek(0)
                df = pd.read_csv(file_obj, encoding=encoding, sep=None, engine='python')
                
                # Vérifier si le DataFrame a du contenu valide
                if not df.empty and len(df.columns) > 1:
                    st.info(f"✅ Encodage détecté: **{encoding}**")
                    return df, encoding
                    
            except (UnicodeDecodeError, pd.errors.ParserError) as e:
                logger.debug(f"Échec avec l'encodage {encoding}: {e}")
                continue
        
        # Si tous les encodages échouent, essayer avec des séparateurs spécifiques
        separators = config.CSV_SEPARATORS
        for separator in separators:
            for encoding in encodings:
                try:
                    file_obj.seek(0)
                    df = pd.read_csv(file_obj, encoding=encoding, sep=separator)
                    
                    if not df.empty and len(df.columns) > 1:
                        st.info(f"✅ Encodage détecté: **{encoding}** avec séparateur **{separator}**")
                        return df, encoding
                        
                except (UnicodeDecodeError, pd.errors.ParserError):
                    continue
        
        # Dernier recours : lire en binaire et essayer de décoder manuellement
        try:
            file_obj.seek(0)
            raw_content = file_obj.read()
            if isinstance(raw_content, str):
                raw_content = raw_content.encode('utf-8')
            
            # Essayer de décoder avec chardet si disponible
            try:
                import chardet
                detected = chardet.detect(raw_content)
                encoding = detected.get('encoding', 'utf-8')
                confidence = detected.get('confidence', 0)
                
                if confidence > 0.7:
                    content = raw_content.decode(encoding)
                    from io import StringIO
                    df = pd.read_csv(StringIO(content), sep=None, engine='python')
                    st.info(f"✅ Encodage détecté avec chardet: **{encoding}** (confiance: {confidence:.2f})")
                    return df, encoding
            except ImportError:
                pass
                
            # Fallback : essayer ISO-8859-1 qui peut décoder pratiquement tout
            content = raw_content.decode('iso-8859-1')
            from io import StringIO
            df = pd.read_csv(StringIO(content), sep=None, engine='python')
            st.warning("⚠️ Encodage par défaut utilisé: **iso-8859-1**")
            return df, 'iso-8859-1'
            
        except Exception as e:
            st.error(f"❌ Impossible de lire le fichier avec aucun encodage testé: {e}")
            return None, None
    
    # ========================= Fonctions utilitaires optimisées =========================
    def normalize_char_for_password(char):
        """Normalise un caractère pour compatibilité Azure AD (supprime accents)"""
        # Mapping des caractères accentués vers non-accentués
        accent_map = {
            'à': 'a', 'á': 'a', 'â': 'a', 'ã': 'a', 'ä': 'a', 'å': 'a',
            'è': 'e', 'é': 'e', 'ê': 'e', 'ë': 'e',
            'ì': 'i', 'í': 'i', 'î': 'i', 'ï': 'i',
            'ò': 'o', 'ó': 'o', 'ô': 'o', 'õ': 'o', 'ö': 'o',
            'ù': 'u', 'ú': 'u', 'û': 'u', 'ü': 'u',
            'ý': 'y', 'ÿ': 'y',
            'ç': 'c', 'ñ': 'n',
            'À': 'a', 'Á': 'a', 'Â': 'a', 'Ã': 'a', 'Ä': 'a', 'Å': 'a',
            'È': 'e', 'É': 'e', 'Ê': 'e', 'Ë': 'e',
            'Ì': 'i', 'Í': 'i', 'Î': 'i', 'Ï': 'i',
            'Ò': 'o', 'Ó': 'o', 'Ô': 'o', 'Õ': 'o', 'Ö': 'o',
            'Ù': 'u', 'Ú': 'u', 'Û': 'u', 'Ü': 'u',
            'Ý': 'y', 'Ÿ': 'y',
            'Ç': 'c', 'Ñ': 'n'
        }
        return accent_map.get(char, char)
    
    def generate_password(firstname="", lastname=""):
        """Génère un mot de passe compatible Azure AD avec préfixe + initiales + chiffres + suffixe
        Format: CFAab1234!*
        - CFA : préfixe
        - ab : première lettre prénom + première lettre nom (minuscules, sans accents)
        - 1234 : 4 chiffres aléatoires
        - !* : suffixe
        """
        # Première lettre du prénom et du nom en minuscule, sans accents
        initial_prenom = normalize_char_for_password(firstname[0]).lower() if firstname else ""
        initial_nom = normalize_char_for_password(lastname[0]).lower() if lastname else ""
        initiales = f"{initial_prenom}{initial_nom}"
        
        # 4 chiffres aléatoires
        chiffres = ''.join(random.choices('0123456789', k=config.PASSWORD_DIGIT_LENGTH))
        
        # Format final: CFA + initiales + chiffres + suffixe
        return f"{config.PASSWORD_PREFIX}{initiales}{chiffres}{config.PASSWORD_SUFFIX}"
    
    def normalize_text_for_username(text):
        """Normalise le texte pour un nom d'utilisateur (supprime accents, caractères spéciaux)"""
        # Mapping des caractères accentués vers non-accentués
        accent_map = {
            'à': 'a', 'á': 'a', 'â': 'a', 'ã': 'a', 'ä': 'a', 'å': 'a',
            'è': 'e', 'é': 'e', 'ê': 'e', 'ë': 'e',
            'ì': 'i', 'í': 'i', 'î': 'i', 'ï': 'i',
            'ò': 'o', 'ó': 'o', 'ô': 'o', 'õ': 'o', 'ö': 'o',
            'ù': 'u', 'ú': 'u', 'û': 'u', 'ü': 'u',
            'ý': 'y', 'ÿ': 'y',
            'ç': 'c', 'ñ': 'n',
            'À': 'a', 'Á': 'a', 'Â': 'a', 'Ã': 'a', 'Ä': 'a', 'Å': 'a',
            'È': 'e', 'É': 'e', 'Ê': 'e', 'Ë': 'e',
            'Ì': 'i', 'Í': 'i', 'Î': 'i', 'Ï': 'i',
            'Ò': 'o', 'Ó': 'o', 'Ô': 'o', 'Õ': 'o', 'Ö': 'o',
            'Ù': 'u', 'Ú': 'u', 'Û': 'u', 'Ü': 'u',
            'Ý': 'y', 'Ÿ': 'y',
            'Ç': 'c', 'Ñ': 'n'
        }
        
        # Normaliser les accents
        normalized = ''.join(accent_map.get(char, char) for char in text)
        
        # Supprimer espaces, apostrophes, tirets et autres caractères spéciaux
        normalized = normalized.replace(" ", "").replace("'", "").replace("-", "").replace("_", "")
        
        # Garder seulement les lettres et chiffres
        normalized = ''.join(char for char in normalized if char.isalnum())
        
        return normalized.lower()
    
    def generate_username(firstname, lastname):
        """Génère un nom d'utilisateur au format prenom.nom
        Exemple: Gaëtan Paviot → gaetan.paviot
                Jean-Baptiste De La Fontaine → jeanbaptiste.delafontaine
        """
        # Normaliser prénom et nom (supprimer accents, espaces, caractères spéciaux)
        prenom_clean = normalize_text_for_username(firstname)
        nom_clean = normalize_text_for_username(lastname)
        
        return f"{prenom_clean}.{nom_clean}"
    
    def get_existing_users():
        """Récupère la liste des utilisateurs depuis la base de données MySQL"""
        try:
            import pymysql
            conn = pymysql.connect(**config.MYSQL_CONFIG)

            query = """
            SELECT
                Login,
                Prénom as 'Prénom',
                Nom,
                Mot_de_passe as 'Mot de passe',
                Classe,
                Groupe,
                Dernière_modification as 'Dernière modification'
            FROM utilisateurs
            WHERE Login IS NOT NULL AND Login != ''
            ORDER BY Nom, Prénom
            """

            df = pd.read_sql(query, conn)
            conn.close()

            return df
        except Exception as e:
            logger.error(f"Erreur lors de la lecture de la base de données MySQL: {e}")
            return pd.DataFrame()

    # ========================= Gestion des mots de passe Excel =========================
    def init_password_excel():
        """Initialise le fichier Excel des mots de passe s'il n'existe pas"""
        if not os.path.exists(PASSWORD_EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Mots de passe"
            
            # Headers
            headers = ["Login", "Prénom", "Nom", "Mot de passe", "Classe", "Groupe", "Dernière modification", "ID"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.fill = header_fill
                
            wb.save(PASSWORD_EXCEL_FILE)
    
    def save_user_to_db(username, firstname, lastname, password, classe="", groupe="Eleves"):
        """Sauvegarde un utilisateur dans la base de données MySQL"""
        # Normaliser la classe avant sauvegarde
        normalized_classe = normalize_class_name(classe) if classe else ""

        # Sauvegarder dans MySQL (table utilisateurs)
        try:
            import pymysql
            conn = pymysql.connect(**config.MYSQL_CONFIG)
            cursor = conn.cursor()

            # UPSERT dans la table utilisateurs
            upsert_query = """
            INSERT INTO utilisateurs (Login, Nom, Prénom, Classe, Groupe, Mot_de_passe, Dernière_modification)
            VALUES (%s, %s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE
                Nom = VALUES(Nom),
                Prénom = VALUES(Prénom),
                Classe = VALUES(Classe),
                Groupe = VALUES(Groupe),
                Mot_de_passe = VALUES(Mot_de_passe),
                Dernière_modification = NOW()
            """

            cursor.execute(upsert_query, (username, lastname, firstname, normalized_classe, groupe, password))
            conn.commit()
            cursor.close()
            conn.close()

            logger.info(f"Utilisateur {username} sauvegardé dans MySQL")
            return True
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde dans MySQL de {username}: {e}")
            return False

    # Alias pour compatibilité avec l'ancien code
    save_user_to_excel = save_user_to_db

    def update_users_created_session(new_users):
        """Met à jour la session avec les nouveaux utilisateurs créés"""
        if os.path.exists(CSV_FILE):
            df_users = pd.read_csv(CSV_FILE)
            st.session_state["users_created"] = df_users.to_dict("records")
        else:
            st.session_state["users_created"] = []

    # ========================= Connexions et commandes SSH =========================
    @contextmanager
    def ssh_connection(server: str, username: str, password: str):
        """Context manager pour les connexions SSH."""
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            client.connect(server, username=username, password=password)
            yield client
        finally:
            client.close()

    def execute_ssh_command(client: paramiko.SSHClient, command: str, password: str) -> Tuple[str, str]:
        """Exécute une commande SSH avec gestion d'erreur."""
        try:
            stdin, stdout, stderr = client.exec_command(command, get_pty=True)
            stdin.write(password + "\n")
            stdin.flush()
            
            output = stdout.read().decode().strip()
            error = stderr.read().decode().strip()
            return output, error
        except Exception as e:
            logger.error(f"Erreur lors de l'exécution de la commande SSH: {e}")
            return "", str(e)

    # ========================= Diagnostics =========================
    def diagnose_environment():
        """Diagnostique l'environnement Kerberos, Samba et Microsoft Graph"""
        st.subheader("🔍 Diagnostic de l'environnement")
        
        diagnostics = {
            f"Keytab existe ({keytab_path})": os.path.exists(keytab_path),
            f"Config Krb5 existe ({krb5_conf})": os.path.exists(krb5_conf),
            f"Script sync existe ({sync_script})": os.path.exists(sync_script),
            f"Config Azure existe ({config_file})": os.path.exists(config_file),
            f"Venv existe ({venv_path})": os.path.exists(venv_path)
        }
        
        # Test de connectivité Samba
        try:
            with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                if client:
                    diagnostics["Connexion Samba réussie"] = True
                else:
                    diagnostics["Connexion Samba réussie"] = False
        except Exception as e:
            diagnostics["Connexion Samba réussie"] = False
            st.error(f"Erreur de connexion Samba: {e}")
        
        # Test de connectivité Microsoft Graph
        try:
            app = ConfidentialClientApplication(
                config.CLIENT_ID,
                authority=f"https://login.microsoftonline.com/{config.TENANT_ID}",
                client_credential=config.CLIENT_SECRET
            )
            token_response = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
            access_token = token_response.get("access_token")
            diagnostics["Connexion Microsoft Graph réussie"] = bool(access_token)
        except Exception as e:
            diagnostics["Connexion Microsoft Graph réussie"] = False
            st.error(f"Erreur de connexion Microsoft Graph: {e}")
        
        for diag, status in diagnostics.items():
            if status:
                st.success(diag)
            else:
                st.error(diag)
        for diag, status in diagnostics.items():
            if status:
                st.success(diag)
            else:
                st.error(diag)

    def test_sync_command():
        """Test de la commande de synchronisation"""
        st.subheader("🧪 Test de la commande de synchronisation")
        
        # Test kinit
        kinit_cmd = f"KRB5_CONFIG={krb5_conf} kinit -k -t {keytab_path} streamlitadmin@CFA-ELEVES.LAN"
        result = subprocess.run(kinit_cmd, shell=True, capture_output=True, text=True, env=env)
        
        st.write("**Test kinit:**")
        st.text("kinit stdout:\n" + result.stdout)
        st.text("kinit stderr:\n" + result.stderr)
        
        # Test klist
        klist_cmd = f"KRB5_CONFIG={krb5_conf} klist"
        result2 = subprocess.run(klist_cmd, shell=True, capture_output=True, text=True, env=env)
        
        st.write("**Test klist:**")
        st.text("klist stdout:\n" + result2.stdout)
        st.text("klist stderr:\n" + result2.stderr)

    # ========================= Affichage résumé avec composants Streamlit =========================
    def display_streamlit_summary(summary_lines):
        """Affiche le résumé avec les composants Streamlit natifs"""
        import re
        
        st.subheader("📊 Résumé de la synchronisation")
        
        # Extraire les données du résumé
        users_sent = 0
        users_errors = 0
        groups_sent = 0
        groups_errors = 0
        passwords_sent = 0 
        passwords_errors = 0
        duration = "N/A"
        is_dry_run = False
        success_rate = 0
        
        for line in summary_lines:
            # Durée
            if "Durée:" in line:
                match = re.search(r'Durée: ([\d.]+) secondes', line)
                if match:
                    duration = f"{match.group(1)}s"
            
            # Mode
            if "Mode: DRY RUN" in line:
                is_dry_run = True
                
            # Utilisateurs
            if "Synchronisés:" in line and "UTILISATEURS" in ''.join(summary_lines[max(0, summary_lines.index(line)-3):summary_lines.index(line)]):
                match = re.search(r'Synchronisés: (\d+)', line)
                if match:
                    users_sent = int(match.group(1))
            if "Erreurs:" in line and "UTILISATEURS" in ''.join(summary_lines[max(0, summary_lines.index(line)-5):summary_lines.index(line)]):
                match = re.search(r'Erreurs: (\d+)', line)
                if match:
                    users_errors = int(match.group(1))
                    
            # Groupes 
            if "Synchronisés:" in line and "GROUPES" in ''.join(summary_lines[max(0, summary_lines.index(line)-3):summary_lines.index(line)]):
                match = re.search(r'Synchronisés: (\d+)', line)
                if match:
                    groups_sent = int(match.group(1))
            if "Erreurs:" in line and "GROUPES" in ''.join(summary_lines[max(0, summary_lines.index(line)-5):summary_lines.index(line)]):
                match = re.search(r'Erreurs: (\d+)', line)
                if match:
                    groups_errors = int(match.group(1))
                    
            # Mots de passe
            if "Hash synchronisés:" in line:
                match = re.search(r'Hash synchronisés: (\d+)', line)
                if match:
                    passwords_sent = int(match.group(1))
            if "Erreurs:" in line and "MOTS DE PASSE" in ''.join(summary_lines[max(0, summary_lines.index(line)-3):summary_lines.index(line)]):
                match = re.search(r'Erreurs: (\d+)', line)
                if match:
                    passwords_errors = int(match.group(1))
                    
            # Taux de réussite
            if "Taux de réussite:" in line:
                match = re.search(r'Taux de réussite: ([\d.]+)%', line)
                if match:
                    success_rate = float(match.group(1))
        
        # Affichage avec composants Streamlit
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric("⏱️ Durée", duration)
            if is_dry_run:
                st.info("🧪 Mode TEST (Dry Run)")
            else:
                st.success("🚀 Mode PRODUCTION")
                
        with col2:
            st.metric("📈 Taux de réussite", f"{success_rate:.1f}%")
            
        # Métriques principales en colonnes
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.subheader("👥 Utilisateurs")
            st.success(f"✅ Synchronisés: **{users_sent}**")
            if users_errors > 0:
                st.error(f"❌ Erreurs: **{users_errors}**")
            else:
                st.success("✅ Aucune erreur")
                
        with col2:
            st.subheader("🗂️ Groupes")  
            st.success(f"✅ Synchronisés: **{groups_sent}**")
            if groups_errors > 0:
                st.error(f"❌ Erreurs: **{groups_errors}**")
            else:
                st.success("✅ Aucune erreur")
                
        with col3:
            st.subheader("🔐 Mots de passe")
            st.success(f"✅ Hash synchronisés: **{passwords_sent}**")
            if passwords_errors > 0:
                st.error(f"❌ Erreurs: **{passwords_errors}**")
            else:
                st.success("✅ Aucune erreur")
        
        # Barre de progression globale
        total_items = users_sent + groups_sent + passwords_sent + users_errors + groups_errors + passwords_errors
        if total_items > 0:
            progress = (users_sent + groups_sent + passwords_sent) / total_items
            st.progress(progress)
            st.caption(f"Progression globale: {users_sent + groups_sent + passwords_sent}/{total_items} éléments traités avec succès")
        
        # Détails dans des expanders
        with st.expander("📋 Détails utilisateurs", expanded=False):
            # Extraire les noms d'utilisateurs du résumé
            user_names = []
            in_user_section = False
            for line in summary_lines:
                if "Premiers utilisateurs synchronisés:" in line:
                    in_user_section = True
                    continue
                elif in_user_section and line.strip().startswith("- "):
                    # Extraire le nom d'utilisateur
                    match = re.search(r'- ([^\s]+)', line.strip())
                    if match:
                        user_names.append(match.group(1))
                elif in_user_section and ("GROUPES" in line or "🗂️" in line):
                    break
                    
            if user_names:
                st.write("**Premiers utilisateurs synchronisés :**")
                for name in user_names[:5]:
                    st.write(f"• {name}")
                if len(user_names) > 5:
                    st.caption(f"... et {users_sent - 5} autres utilisateurs")
            else:
                st.info("Aucun détail utilisateur disponible")
        
        # Statut final
        if success_rate == 100.0:
            st.balloons()
            st.success("🎉 Synchronisation terminée avec succès - Aucune erreur !")
        elif success_rate >= 90.0:
            st.success(f"✅ Synchronisation terminée avec {success_rate:.1f}% de réussite")
        else:
            st.warning(f"⚠️ Synchronisation terminée avec des erreurs ({success_rate:.1f}% de réussite)")

    # ========================= Synchronisation =========================
    def run_sync_with_live_logs(is_dryrun=True, timeout_seconds=120, use_summary=True):
        """Exécute la synchronisation Azure AD avec logs en temps réel"""
        import subprocess
        import threading
        import queue
        import time
        
        # Créer les conteneurs pour affichage en temps réel
        log_container = st.empty()
        output_lines = []
        error_lines = []
        
        try:
            # Vérifier si un processus de sync est déjà en cours
            check_cmd = ["pgrep", "-f", "run_sync.py"]
            check_result = subprocess.run(check_cmd, capture_output=True, text=True)
            
            if check_result.returncode == 0 and check_result.stdout.strip():
                st.warning("⚠️ Une synchronisation est déjà en cours. Attendez qu'elle se termine ou redémarrez l'application.")
                # Tuer les anciens processus
                subprocess.run(["pkill", "-f", "run_sync.py"], capture_output=True)
                time.sleep(2)
            
            # Initialiser le ticket Kerberos avant la synchronisation
            st.info("🔐 Initialisation de l'authentification Kerberos...")
            kinit_cmd = ["kinit", "-k", "-t", keytab_path, "streamlitadmin@CFA-ELEVES.LAN"]
            kinit_result = subprocess.run(kinit_cmd, env=env, capture_output=True, text=True)
            
            if kinit_result.returncode != 0:
                st.error(f"❌ Échec de l'authentification Kerberos: {kinit_result.stderr}")
                return None
            else:
                st.success("✅ Authentification Kerberos réussie")
            
            # Choisir le script selon l'option
            script_to_use = config.SYNC_SCRIPT_WITH_SUMMARY if use_summary else sync_script
            cmd = [f"{venv_path}/bin/python", script_to_use]
            if is_dryrun:
                cmd.append("--dryrun")
            
            # Démarrer le processus
            process = subprocess.Popen(
                cmd,
                cwd="/home/samba-sync-ad",
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,  # Rediriger stderr vers stdout
                text=True,
                bufsize=0,  # Unbuffered
                universal_newlines=True,
                preexec_fn=lambda: None  # Force immediate output
            )
            
            start_time = time.time()
            
            # Lire les logs ligne par ligne en temps réel
            import select
            while True:
                # Vérifier si le processus est terminé
                if process.poll() is not None:
                    # Lire les dernières lignes
                    remaining_output = process.stdout.read()
                    if remaining_output:
                        output_lines.extend(remaining_output.strip().split('\n'))
                    break
                    
                # Vérifier le timeout
                if time.time() - start_time > timeout_seconds:
                    process.terminate()
                    st.error(f"❌ Timeout après {timeout_seconds} secondes")
                    return None
                
                # Utiliser select pour lire de manière non-bloquante
                ready, _, _ = select.select([process.stdout], [], [], 0.1)
                if ready:
                    line = process.stdout.readline()
                    if line:
                        output_lines.append(line.rstrip())
                        
                        # Mettre à jour l'affichage plus fréquemment
                        if len(output_lines) % 2 == 0 or len(output_lines) > 20:
                            with log_container.container():
                                st.subheader("📋 Logs de synchronisation en temps réel")
                                # Afficher les 30 dernières lignes
                                recent_logs = output_lines[-30:] if len(output_lines) > 30 else output_lines
                                st.code('\n'.join(recent_logs), language="text")
                
                time.sleep(0.1)  # Petite pause
            
            # Lire les dernières lignes
            remaining_output = process.stdout.read()
            if remaining_output:
                output_lines.extend(remaining_output.strip().split('\n'))
            
            # Affichage final
            with log_container.container():
                if use_summary and output_lines:
                    # Séparer le résumé des logs JSON
                    summary_start = -1
                    for i, line in enumerate(output_lines):
                        if "📊 RÉSUMÉ DE LA SYNCHRONISATION AZURE AD" in line:
                            summary_start = i
                            break
                    
                    if summary_start >= 0:
                        # Afficher le résumé en premier
                        summary_lines = output_lines[summary_start:]
                        json_logs = output_lines[:summary_start]
                        
                        # Extraire les données du résumé pour un affichage Streamlit
                        display_streamlit_summary(summary_lines)
                        
                        # Afficher les logs JSON dans un expander fermé
                        if json_logs:
                            with st.expander("🔍 Voir les logs JSON détaillés", expanded=False):
                                st.code('\n'.join(json_logs), language="json")
                    else:
                        st.subheader("📋 Logs de synchronisation")
                        st.code('\n'.join(output_lines), language="text")
                else:
                    st.subheader("📋 Logs complets de synchronisation")
                    st.code('\n'.join(output_lines), language="text")
            
            # Créer un objet résultat compatible
            class SyncResult:
                def __init__(self, returncode, stdout):
                    self.returncode = returncode
                    self.stdout = stdout
                    self.stderr = ""
            
            return SyncResult(process.returncode, '\n'.join(output_lines))
            
        except Exception as e:
            st.error(f"❌ Erreur lors de l'exécution : {e}")
            return None

    # ========================= Interface utilisateur en onglets =========================
    
    st.title(config.APP_TITLE)
    
    # ========================= ONGLETS PRINCIPAUX =========================
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "👤 Création utilisateur",
        "📂 Import CSV",
        "🔄 Synchronisation AD",
        "✏️ Modifier mot de passe",
        "🗑️ Suppression",
        "👥 Groupes",
        "🔄 Sync IMFR/SAMBA",
        "🔧 Outils"
    ])
    
    # ========================= ONGLET 1: CRÉATION UTILISATEUR =========================
    with tab1:
        st.header("Création d'utilisateur Samba")
        
        col1, col2 = st.columns(2)
        
        with col1:
            firstname = st.text_input("Prénom", key="create_firstname")
            lastname = st.text_input("Nom", key="create_lastname")
            password = st.text_input("Mot de passe (laisser vide pour générer automatiquement)", type="password", key="create_password")
            
        with col2:
            classe_input = st.text_input("Classe (optionnel)", key="create_classe")
            # Normaliser la classe saisie
            classe = normalize_class_name(classe_input) if classe_input else ""
            
            # Afficher un aperçu de la classe normalisée
            if classe_input and classe_input != classe:
                st.info(f"Classe normalisée: **{classe_input}** → **{classe}**")
                
            # Choix du groupe cible
            groupe_cible = st.radio(
                "Choisir le groupe :",
                ["Eleves", "formateur"],
                index=0,
                horizontal=True,
                key="create_groupe"
            )
            
            # Attribution des licences Microsoft 365 réelles depuis Entra ID
            st.subheader("🔑 Attribution des licences Microsoft 365")
            
            # Récupération des licences disponibles depuis Microsoft Graph
            with st.spinner("📡 Récupération des licences disponibles depuis Entra ID..."):
                available_licenses = get_available_licenses()
            
            if available_licenses:
                st.success(f"✅ {len(available_licenses)} types de licences trouvés dans Entra ID")
                
                # Interface de sélection des licences
                selected_licenses = []
                
                # Interface de sélection des licences - sans colonnes imbriquées
                assign_student_license = st.checkbox(
                    f"🎓 Attribuer une licence étudiant ({config.LICENSE_GROUP_STUDENTS})",
                    value=False,  # Par défaut désactivé pour éviter l'attribution automatique
                    help=f"Ajoute l'utilisateur au groupe de licence (ID: {config.LICENSE_GROUP_STUDENTS_ID})",
                    key="student_license"
                )
                if assign_student_license:
                    selected_licenses.append(config.LICENSE_GROUP_STUDENTS)
                

                # Affichage des groupes sélectionnés
                if selected_licenses:
                    st.info("💡 Les licences seront attribuées via ces groupes lors de la synchronisation Azure AD")
                    st.write(f"**Groupes de licences sélectionnés:** {', '.join(selected_licenses)}")
                    
            else:
                st.warning("⚠️ Impossible de récupérer les licences depuis Entra ID")
                st.info("📝 Configuration manuelle des groupes de licences")
                
                selected_licenses = []
                # Interface de fallback - sans colonnes imbriquées
                assign_student_license = st.checkbox(
                    f"🎓 Attribuer une licence étudiant ({config.LICENSE_GROUP_STUDENTS})",
                    value=False,  # Par défaut désactivé pour éviter l'attribution automatique
                    key="fallback_student_license"
                )
                assign_teacher_license = st.checkbox(
                    f"👨‍🏫 Attribuer une licence formateur ({config.LICENSE_GROUP_TEACHERS})",
                    value=False,  # Par défaut désactivé pour éviter l'attribution automatique
                    key="fallback_teacher_license"
                )
                
                if assign_student_license:
                    selected_licenses.append(config.LICENSE_GROUP_STUDENTS)
                if assign_teacher_license:
                    selected_licenses.append(config.LICENSE_GROUP_TEACHERS)
            if st.button("Créer l'utilisateur Samba", type="primary"):
                if not firstname or not lastname:
                    st.error("Merci de remplir tous les champs obligatoires.")
                else:
                    # Génération automatique du username
                    username = generate_username(firstname, lastname)
                    
                    # Si aucun mot de passe fourni, générer automatiquement
                    if not password:
                        password = generate_password(firstname, lastname)
                    
                    # Créer la commande Samba-Tool pour créer l'utilisateur
                    cmd_create = f"sudo -S samba-tool user create {username} {password} " \
                        f"--description='{classe}'"

                    # Commandes pour les groupes
                    group_commands = []
                    all_groups = [groupe_cible]
                    
                    # Ajout au groupe principal
                    group_commands.append(f"sudo -S samba-tool group addmembers {groupe_cible} {username}")
                    
                    # Ajout aux groupes de licences sélectionnés
                    if 'selected_licenses' in locals() and selected_licenses:
                        for license_group in selected_licenses:
                            # Créer le groupe s'il n'existe pas
                            group_commands.append(f"sudo -S samba-tool group add {license_group} || true")  # || true pour ignorer si existe déjà
                            # Ajouter l'utilisateur au groupe
                            group_commands.append(f"sudo -S samba-tool group addmembers {license_group} {username}")
                            all_groups.append(license_group)

                    with st.expander("Détails de création", expanded=True):
                        st.write(f"**Utilisateur:** {username}")
                        st.write(f"**Mot de passe généré:** {password}")
                        st.write(f"**Groupe principal:** {groupe_cible}")
                        st.write(f"**Classe:** {classe}")
                        if len(all_groups) > 1:
                            st.write(f"**Groupes de licences:** {', '.join(all_groups[1:])}")
                            st.info("🔄 Les licences seront activées automatiquement lors de la prochaine synchronisation Azure AD")
                    
                    try:
                        with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                            if client:
                                # Création de l'utilisateur
                                output_create, error_create = execute_ssh_command(client, cmd_create, config.SAMBA_PWD)
                                if error_create:
                                    st.error(f"❌ Erreur lors de la création de {username}: {error_create}")
                                else:
                                    # Ajout aux groupes (principal + licences)
                                    all_outputs = [output_create]
                                    group_success = True
                                    group_errors = []
                                    
                                    for idx, group_cmd in enumerate(group_commands):
                                        output_group, error_group = execute_ssh_command(client, group_cmd, config.SAMBA_PWD)
                                        all_outputs.append(output_group)
                                        
                                        if error_group and "already exists" not in error_group.lower():
                                            group_success = False
                                            group_name = all_groups[min(idx, len(all_groups)-1)]
                                            group_errors.append(f"Groupe '{group_name}': {error_group}")
                                    
                                    if group_errors:
                                        st.warning(f"⚠️ Utilisateur créé ({username}) mais problèmes avec certains groupes:")
                                        for error in group_errors:
                                            st.warning(f"  • {error}")
                                    else:
                                        st.success(f"✅ Utilisateur {firstname} {lastname} créé avec login **{username}**")
                                        if len(all_groups) > 1:
                                            st.success(f"✅ Ajouté aux groupes: {', '.join(all_groups)}")
                                    
                                    # Sauvegarder dans Excel
                                    save_user_to_excel(username, firstname, lastname, password, classe, groupe_cible)
                                    update_users_created_session([{"username": username, "firstname": firstname, "lastname": lastname, "password": password, "classe": classe, "groupe": groupe_cible}])

                                    with st.expander("Logs de création", expanded=False):
                                        st.text("\n".join(all_outputs))
                            else:
                                st.error("❌ Impossible de se connecter au serveur Samba")
                                # Sauvegarder quand même dans l'Excel (pour traçabilité)
                                save_user_to_excel(username, firstname, lastname, password, classe, groupe_cible)

                    except Exception as e:
                        st.error(f"❌ Erreur : {str(e)}")

    # ========================= ONGLET 2: IMPORT CSV =========================
    with tab2:
        st.header("Import CSV - Création en masse")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.info("Format CSV requis: **nom,prenom,classe,groupe**")
            
            # Exemple de fichier CSV  
            example_data = {
                "nom": ["Dupont", "Martin"],
                "prenom": ["Jean", "Marie"], 
                "classe": ["T Bac 2", "BTS 1"],
                "groupe": ["Eleves", "Eleves"]
            }
            st.download_button(
                label="📥 Télécharger un exemple CSV",
                data=pd.DataFrame(example_data).to_csv(index=False),
                file_name="exemple_import.csv",
                mime="text/csv"
            )
        
        with col2:
            uploaded_file = st.file_uploader("Choisir un fichier CSV", type="csv")
            
        if uploaded_file is not None:
            try:
                # Détecter automatiquement l'encodage et lire le CSV
                df, detected_encoding = detect_csv_encoding(uploaded_file)
                
                if df is None:
                    st.error("❌ Impossible de lire le fichier CSV. Vérifiez le format et l'encodage.")
                    st.info(config.HELP_MESSAGES['csv_format'])
                    st.info(config.HELP_MESSAGES['encodings'])
                    return
                
                st.write("Aperçu du fichier:", df.head())
                
                # Normaliser les noms de colonnes (gérer les variations possibles)
                column_mapping = config.CSV_COLUMN_MAPPING
                
                # Normaliser les noms de colonnes
                df.columns = df.columns.str.lower().str.strip()
                for old_name, new_name in column_mapping.items():
                    if old_name in df.columns:
                        df = df.rename(columns={old_name: new_name})
                
                # Si les colonnes n'ont toujours pas été renommées, essayer une approche plus flexible
                if 'prenom' not in df.columns:
                    # Chercher toute colonne contenant "prénom" ou "prenom"
                    for col in df.columns:
                        if 'prénom' in col.lower() or 'prenom' in col.lower():
                            df = df.rename(columns={col: 'prenom'})
                            st.info(f"🔄 Colonne '{col}' renommée en 'prenom'")
                            break
                
                if 'nom' not in df.columns:
                    # Chercher toute colonne contenant "nom" mais pas "prénom"
                    for col in df.columns:
                        if 'nom' in col.lower() and 'prénom' not in col.lower() and 'prenom' not in col.lower():
                            df = df.rename(columns={col: 'nom'})
                            st.info(f"🔄 Colonne '{col}' renommée en 'nom'")
                            break
                
                if 'classe' not in df.columns:
                    # Chercher toute colonne contenant "formation", "classe" ou "parcours"
                    for col in df.columns:
                        if any(word in col.lower() for word in ['formation', 'classe', 'parcours']):
                            df = df.rename(columns={col: 'classe'})
                            st.info(f"🔄 Colonne '{col}' renommée en 'classe'")
                            break
                
                # Vérification des colonnes requises
                required_cols = ['nom', 'prenom', 'classe']
                missing_cols = [col for col in required_cols if col not in df.columns]
                
                if missing_cols:
                    st.error(f"❌ Colonnes manquantes: {', '.join(missing_cols)}")
                else:
                    st.success(f"✅ Fichier valide - {len(df)} utilisateurs à traiter")
                    
                    # ========================= CONFIGURATION D'IMPORT =========================
                    st.subheader("⚙️ Configuration de l'import")
                    col_config1, col_config2, col_config3 = st.columns(3)
                    # Place les widgets directement dans les colonnes, sans bloc 'with'
                    groupe_par_defaut = col_config1.radio(
                        "Groupe par défaut pour tous les utilisateurs:",
                        ["Eleves", "formateur"],
                        index=0,
                        key="groupe_defaut_csv"
                    )
                    show_preview = col_config2.checkbox("Afficher l'aperçu détaillé", value=True, key="show_preview")
                    simulate_passwords = col_config2.checkbox("Simuler les mots de passe", value=True, key="simulate_passwords")
                    dry_run_csv = col_config3.checkbox("Mode simulation uniquement", value=False, key="csv_dry_run")
                    if dry_run_csv:
                        col_config3.info("🔍 Mode simulation : Aucun compte ne sera créé")
                    else:
                        col_config3.warning("⚠️ Mode création réelle activé")
                    
                    # ========================= SÉLECTION DES LICENCES =========================
                    st.subheader("🔑 Attribution des licences Microsoft 365")
                    st.info("ℹ️ Cette option s'appliquera à tous les utilisateurs de l'import CSV")
                    
                    csv_selected_licenses = []
                    
                    # Récupération des licences disponibles depuis Microsoft Graph
                    with st.spinner("📡 Récupération des licences disponibles..."):
                        available_licenses = get_available_licenses()
                    
                    if available_licenses:
                        st.success(f"✅ {len(available_licenses)} types de licences trouvés dans Entra ID")
                        
                        # Interface de sélection des licences pour import en masse
                        col_lic1, col_lic2 = st.columns(2)
                        
                        with col_lic1:
                            csv_assign_student_license = st.checkbox(
                                f"🎓 Attribuer licence étudiant ({config.LICENSE_GROUP_STUDENTS})",
                                value=False,  # Par défaut désactivé pour éviter l'attribution automatique
                                help=f"Ajoute tous les utilisateurs au groupe (ID: {config.LICENSE_GROUP_STUDENTS_ID})",
                                key="csv_student_license"
                            )
                            if csv_assign_student_license:
                                csv_selected_licenses.append(config.LICENSE_GROUP_STUDENTS)
                        
                        with col_lic2:
                            csv_assign_teacher_license = st.checkbox(
                                f"👨‍🏫 Attribuer licence formateur ({config.LICENSE_GROUP_TEACHERS})",
                                value=False,  # Par défaut désactivé pour éviter l'attribution automatique
                                help="Ajoute tous les utilisateurs au groupe formateurs",
                                key="csv_teacher_license"
                            )
                            if csv_assign_teacher_license:
                                csv_selected_licenses.append(config.LICENSE_GROUP_TEACHERS)
                        
                        # Affichage des licences disponibles dans Entra ID (informationnel)
                        with st.expander("📋 Licences disponibles dans Entra ID"):
                            for sku_id, license_info in available_licenses.items():
                                st.write(f"**{license_info['displayName']}** ({license_info['partNumber']})")
                                st.write(f"   • Unités consommées: {license_info['consumedUnits']}")
                                st.write(f"   • Unités activées: {license_info['enabledUnits']}")
                                st.write(f"   • Unités disponibles: {license_info['availableUnits']}")
                                st.write("---")
                        
                    else:
                        st.warning("⚠️ Impossible de récupérer les licences depuis Entra ID")
                        st.info("📝 Configuration manuelle des groupes de licences")
                        
                        # Interface de fallback
                        csv_assign_student_license = st.checkbox(
                            f"🎓 Attribuer licence étudiant ({config.LICENSE_GROUP_STUDENTS})",
                            value=False,  # Par défaut désactivé pour éviter l'attribution automatique
                            key="csv_fallback_student_license"
                        )
                        csv_assign_teacher_license = st.checkbox(
                            f"👨‍🏫 Attribuer licence formateur ({config.LICENSE_GROUP_TEACHERS})",
                            value=False,  # Par défaut désactivé pour éviter l'attribution automatique
                            key="csv_fallback_teacher_license"
                        )
                        
                        if csv_assign_student_license:
                            csv_selected_licenses.append(config.LICENSE_GROUP_STUDENTS)
                        if csv_assign_teacher_license:
                            csv_selected_licenses.append(config.LICENSE_GROUP_TEACHERS)
                    
                    # Affichage des groupes sélectionnés
                    if csv_selected_licenses:
                        st.info("💡 Les licences seront attribuées via ces groupes lors de la synchronisation Azure AD")
                        st.write(f"**Groupes de licences sélectionnés:** {', '.join(csv_selected_licenses)}")
                    
                    # ========================= APERÇU DES COMPTES =========================
                    if show_preview:
                        st.subheader("👁️ Aperçu des comptes à créer")
                        
                        # Stockage des données d'aperçu dans session_state pour cohérence
                        if 'csv_preview_data' not in st.session_state or st.button("🔄 Actualiser l'aperçu", key="refresh_preview"):
                            st.session_state.csv_preview_data = []
                            
                            for idx, row in df.iterrows():
                                firstname = str(row['prenom']).strip()
                                lastname = str(row['nom']).strip()
                                classe = str(row['classe']).strip()
                                # Normaliser la classe pour l'aperçu
                                classe_normalized = normalize_class_name(classe) if classe and classe != 'nan' else ""
                                
                                # Génération du username et password pour l'aperçu
                                username = generate_username(firstname, lastname)
                                password = generate_password(firstname, lastname) if simulate_passwords else "CFA****"
                                
                                # Construire la liste des groupes (principal + licences)
                                groups_list = [groupe_par_defaut]
                                if 'csv_selected_licenses' in locals() and csv_selected_licenses:
                                    groups_list.extend(csv_selected_licenses)
                                
                                st.session_state.csv_preview_data.append({
                                    "Prénom": firstname,
                                    "Nom": lastname,
                                    "Login généré": username,
                                    "Mot de passe": password,
                                    "Classe originale": classe,
                                    "Classe normalisée": classe_normalized,
                                    "Groupe principal": groupe_par_defaut,
                                    "Groupes licences": ", ".join(csv_selected_licenses) if csv_selected_licenses else "Aucune"
                                })
                        
                        df_preview = pd.DataFrame(st.session_state.csv_preview_data)
                        
                        # Affichage avec possibilité de télécharger l'aperçu
                        st.dataframe(df_preview, hide_index=True, use_container_width=True)
                        
                        # Statistiques de l'aperçu
                        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                        with col_stat1:
                            st.metric("Total comptes", len(df_preview))
                        with col_stat2:
                            unique_classes = df_preview['Classe normalisée'].unique()
                            st.metric("Classes différentes", len([c for c in unique_classes if c]))
                        with col_stat3:
                            st.metric("Groupe", groupe_par_defaut)
                        with col_stat4:
                            if simulate_passwords:
                                unique_passwords = len(set(df_preview['Mot de passe']))
                                st.metric("Mots de passe uniques", unique_passwords)
                            else:
                                st.metric("Mots de passe", "Simulés")
                        
                        # Bouton de téléchargement de l'aperçu
                        csv_preview = df_preview.to_csv(index=False, encoding='utf-8-sig')
                        st.download_button(
                            label="📥 Télécharger l'aperçu complet CSV",
                            data=csv_preview,
                            file_name="apercu_comptes_a_creer.csv",
                            mime="text/csv"
                        )
                        
                        # Avertissements sur les classes non reconnues
                        classes_non_reconnues = []
                        for idx, row in df.iterrows():
                            classe_orig = str(row['classe']).strip()
                            if classe_orig and classe_orig != 'nan' and classe_orig not in CLASS_MAPPING:
                                if classe_orig not in classes_non_reconnues:
                                    classes_non_reconnues.append(classe_orig)
                        
                        if classes_non_reconnues:
                            st.warning(f"⚠️ Classes non reconnues dans le mapping : {', '.join(classes_non_reconnues)}")
                            st.info("Ces classes seront normalisées automatiquement selon les règles de mapping.")
                    
                    # ========================= BOUTONS D'ACTION =========================
                    st.markdown("---")
                    col_btn1, col_btn2 = st.columns(2)
                    
                    with col_btn1:
                        if dry_run_csv:
                            simulation_btn = st.button("🔍 Lancer la simulation", type="primary", key="simulation_btn")
                        else:
                            st.info("💡 Conseil: Activez d'abord le mode simulation pour tester")
                    
                    with col_btn2:
                        if not dry_run_csv:
                            creation_btn = st.button("🚀 CRÉER LES COMPTES", type="primary", key="creation_btn")
                            if creation_btn:
                                st.warning("⚠️ Création réelle des comptes en cours...")
                        else:
                            st.info("Désactivez le mode simulation pour créer")
                    
                    # ========================= TRAITEMENT =========================
                    process_csv = (dry_run_csv and 'simulation_btn' in locals() and simulation_btn) or \
                                 (not dry_run_csv and 'creation_btn' in locals() and creation_btn)
                    
                    if process_csv:
                        # ========================= INITIALISATION =========================
                        st.subheader("🔄 Traitement en cours")
                        
                        # Conteneurs d'affichage
                        progress_container = st.container()
                        logs_container = st.container()
                        summary_container = st.container()
                        
                        with progress_container:
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            progress_text = st.empty()
                            
                        with logs_container:
                            st.subheader("📋 Logs détaillés")
                            log_placeholder = st.empty()
                            
                        # Variables de suivi
                        results = []
                        detailed_logs = []
                        success_count = 0
                        error_count = 0
                        warning_count = 0
                        
                        # ========================= TRAITEMENT =========================
                        for idx, row in df.iterrows():
                            firstname = str(row['prenom']).strip()
                            lastname = str(row['nom']).strip()
                            classe = str(row['classe']).strip()
                            # Normaliser la classe
                            classe = normalize_class_name(classe) if classe and classe != 'nan' else ""
                            
                            # Génération automatique du username et password
                            username = generate_username(firstname, lastname)
                            password = generate_password(firstname, lastname)
                            
                            # Mise à jour du statut
                            current_progress = (idx + 1) / len(df)
                            progress_bar.progress(current_progress)
                            status_text.text(f"🔄 Traitement: {firstname} {lastname} ({idx+1}/{len(df)})")
                            progress_text.text(f"Progression: {int(current_progress * 100)}%")
                            
                            # Log de début
                            log_entry = f"🔄 **{idx+1}.** Traitement de **{firstname} {lastname}**\\n"
                            log_entry += f"   • Login: `{username}`\\n"
                            log_entry += f"   • Classe: `{classe}`\\n"
                            log_entry += f"   • Groupe: `{groupe_par_defaut}`\\n"
                            
                            # Commandes Samba-Tool
                            cmd_create = f"sudo -S samba-tool user create {username} {password} " \
                                f"--description='{classe}'"
                            
                            # Commandes pour les groupes (principal + licences)
                            group_commands = []
                            all_groups = [groupe_par_defaut]
                            
                            # Ajout au groupe principal
                            group_commands.append(f"sudo -S samba-tool group addmembers {groupe_par_defaut} {username}")
                            
                            # Ajout aux groupes de licences sélectionnés (CSV)
                            if 'csv_selected_licenses' in locals() and csv_selected_licenses:
                                for license_group in csv_selected_licenses:
                                    # Créer le groupe s'il n'existe pas
                                    group_commands.append(f"sudo -S samba-tool group add {license_group} || true")
                                    # Ajouter l'utilisateur au groupe
                                    group_commands.append(f"sudo -S samba-tool group addmembers {license_group} {username}")
                                    all_groups.append(license_group)
                                
                                log_entry += f"   • Groupes de licences: `{', '.join(csv_selected_licenses)}`\\n"
                            
                            if dry_run_csv:
                                # Mode simulation - PAS DE SAUVEGARDE EN DRY RUN
                                log_entry += f"   • **SIMULATION** - Aucune création réelle\\n"
                                log_entry += f"   • Mot de passe généré: `{password}`\\n"
                                log_entry += f"   • ✅ **Simulation réussie (pas de sauvegarde)**\\n"
                                
                                results.append(f"🔍 {firstname} {lastname} ({username}) - SIMULATION")
                                # NOTE: En mode dry run, on ne sauvegarde PAS dans Excel
                                success_count += 1
                                
                            else:
                                # Mode création réelle
                                log_entry += f"   • Création du compte sur Samba...\\n"
                                
                                try:
                                    with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                                        if client:
                                            # Création utilisateur
                                            log_entry += f"   • Exécution: `samba-tool user create {username}...`\\n"
                                            output_create, error_create = execute_ssh_command(client, cmd_create, config.SAMBA_PWD)
                                            
                                            if error_create:
                                                log_entry += f"   • ❌ **Erreur création**: {error_create[:100]}...\\n"
                                                results.append(f"❌ {firstname} {lastname}: {error_create}")
                                                error_count += 1
                                            else:
                                                log_entry += f"   • ✅ Utilisateur créé avec succès\\n"
                                                
                                                # Ajout aux groupes (principal + licences)
                                                group_errors = []
                                                for group_cmd in group_commands:
                                                    log_entry += f"   • Exécution: `{group_cmd.split('sudo -S ')[1]}`...\\n"
                                                    output_group, error_group = execute_ssh_command(client, group_cmd, config.SAMBA_PWD)
                                                    
                                                    if error_group and "already exists" not in error_group.lower():
                                                        group_errors.append(error_group[:50])
                                                
                                                if group_errors:
                                                    log_entry += f"   • ⚠️ **Avertissements**: Erreurs groupes: {'; '.join(group_errors)}...\\n"
                                                    results.append(f"⚠️ {firstname} {lastname}: Créé mais erreur groupes")
                                                    warning_count += 1
                                                else:
                                                    log_entry += f"   • ✅ Ajouté au groupe avec succès\\n"
                                                    log_entry += f"   • 🔑 Mot de passe: `{password}`\\n"
                                                    results.append(f"✅ {firstname} {lastname} ({username})")
                                                    success_count += 1
                                                
                                                # Sauvegarder dans Excel
                                                save_user_to_excel(username, firstname, lastname, password, classe, groupe_par_defaut)
                                                log_entry += f"   • 💾 Sauvegardé dans Excel\\n"
                                        else:
                                            log_entry += f"   • ❌ **Erreur**: Impossible de se connecter au serveur Samba\\n"
                                            results.append(f"❌ {firstname} {lastname}: Connexion impossible")
                                            save_user_to_excel(username, firstname, lastname, password, classe, groupe_par_defaut)
                                            error_count += 1
                                            
                                except Exception as e:
                                    log_entry += f"   • ❌ **Exception**: {str(e)}\\n"
                                    results.append(f"❌ {firstname} {lastname}: {str(e)}")
                                    error_count += 1
                            
                            log_entry += f"\\n---\\n"
                            detailed_logs.append(log_entry)
                            
                            # Mise à jour des logs en temps réel
                            with log_placeholder:
                                if len(detailed_logs) <= 5:
                                    # Afficher tous les logs si peu d'utilisateurs
                                    for log in detailed_logs:
                                        st.markdown(log)
                                else:
                                    # Afficher seulement les 3 derniers logs si beaucoup d'utilisateurs
                                    st.markdown("*... (logs précédents masqués pour la performance)*")
                                    for log in detailed_logs[-3:]:
                                        st.markdown(log)
                        
                        # ========================= FINALISATION =========================
                        status_text.text("✅ Traitement terminé !")
                        progress_text.text("Terminé à 100%")
                        
                        # Résumé final
                        with summary_container:
                            st.subheader("📊 Résumé du traitement")
                            
                            col_s1, col_s2, col_s3, col_s4 = st.columns(4)
                            with col_s1:
                                st.metric("Total traité", len(df))
                            with col_s2:
                                st.metric("Succès", success_count, delta=f"{int(success_count/len(df)*100)}%")
                            with col_s3:
                                st.metric("Avertissements", warning_count)
                            with col_s4:
                                st.metric("Erreurs", error_count)
                            
                            if dry_run_csv:
                                st.success(f"🔍 **Simulation terminée** - {len(results)} comptes analysés")
                                st.info("💡 Désactivez le mode simulation et cliquez sur 'CRÉER LES COMPTES' pour la création réelle.")
                            else:
                                if error_count == 0 and warning_count == 0:
                                    st.success(f"🎉 **Tous les comptes ont été créés avec succès !** ({success_count}/{len(df)})")
                                elif error_count == 0:
                                    st.warning(f"⚠️ **Création terminée avec {warning_count} avertissements** ({success_count}/{len(df)} réussis)")
                                else:
                                    st.error(f"❌ **Création terminée avec {error_count} erreurs** ({success_count}/{len(df)} réussis)")
                            
                            # Logs complets téléchargeables
                            full_logs = "\\n".join(detailed_logs)
                            st.download_button(
                                label="📥 Télécharger les logs complets",
                                data=full_logs,
                                file_name=f"logs_creation_{'simulation' if dry_run_csv else 'reel'}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.txt",
                                mime="text/plain"
                            )
                                
            except Exception as e:
                st.error(f"Erreur lors de la lecture du fichier CSV: {str(e)}")

    # ========================= ONGLET 3: SYNCHRONISATION AD =========================
    with tab3:
        st.header("Synchronisation vers Azure AD")
        
        col1, col2 = st.columns(2)
        
        with col1:
            timeout_config = st.slider("Timeout (secondes)", 10, 600, 60, 10)
            st.info("💡 **Temps recommandés (mis à jour):**\n- Dry run: 10-30s\n- Sync réelle: 30-60s\n- Si lenteur: 120-180s")
            dryrun = st.checkbox("Dry Run (mode test)", value=True)
            
        with col2:
            st.info(config.HELP_MESSAGES['dry_run'])
            use_summary = st.checkbox("📊 Affichage résumé (plus lisible)", value=True)
            st.caption("✅ Résumé : affichage digestible\n❌ Logs bruts : JSON détaillé")
            
        if st.button("🔄 Exécuter la synchronisation AD", type="primary"):
            st.subheader("🔄 Synchronisation en cours...")
            
            # Indicateur de progression
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            status_text.text("🚀 Initialisation de la synchronisation...")
            progress_bar.progress(10)
            
            import time
            time.sleep(0.5)
            
            status_text.text(f"📡 Exécution du script (timeout: {timeout_config}s)...")
            progress_bar.progress(20)
            
            # Utiliser la version avec résumé ou logs selon le choix
            result = run_sync_with_live_logs(is_dryrun=dryrun, timeout_seconds=timeout_config, use_summary=use_summary)
            progress_bar.progress(90)
                
            status_text.text("✅ Traitement terminé !")
            progress_bar.progress(100)
                
            # Affichage des résultats avec logs visibles par défaut
            if result and result.returncode == 0:
                st.success("✅ Synchronisation terminée avec succès")
                
                # Logs toujours visibles pour débugger les lenteurs
                st.subheader("📋 Logs détaillés de synchronisation")
                if result.stdout:
                    st.code(result.stdout, language="text")
                else:
                    st.info("Aucun log de sortie disponible")
                    
            elif result:
                st.error("❌ Erreur lors de la synchronisation")
                
                st.subheader("📋 Logs d'erreur")
                if result.stderr:
                    st.error("**Erreurs:**")
                    st.code(result.stderr, language="text")
                    
                st.subheader("📋 Logs de sortie")
                if result.stdout:
                    st.code(result.stdout, language="text")
                else:
                    st.info("Aucun log de sortie disponible")
                    
            else:
                st.error("❌ Le processus n'a pas pu se terminer correctement")
                st.warning(f"⏰ Timeout après {timeout_config} secondes - Le script prend trop de temps")
    # ========================= ONGLET 4: MODIFIER MOT DE PASSE =========================
    with tab4:
        st.header("Modifier le mot de passe d'un utilisateur")
        
        # Récupérer la liste des utilisateurs existants
        df_users = get_existing_users()
        
        if not df_users.empty and all(col in df_users.columns for col in ['Prénom', 'Nom', 'Login']):
            col1, col2 = st.columns(2)
            
            with col1:
                # Sélection de l'utilisateur
                user_options = [f"{row['Prénom']} {row['Nom']} ({row['Login']})" for _, row in df_users.iterrows()]
                selected_user = st.selectbox("Sélectionner un utilisateur:", user_options, key="modify_user_select")
                
                if selected_user:
                    # Extraire les infos de l'utilisateur sélectionné
                    selected_index = user_options.index(selected_user)
                    user_data = df_users.iloc[selected_index]
                    
                    st.info(f"**Utilisateur sélectionné:** {user_data['Prénom']} {user_data['Nom']}")
                    st.info(f"**Login:** {user_data['Login']}")
                    st.info(f"**Classe:** {user_data.get('Classe', 'Non définie')}")
                    
            with col2:
                # Saisie du nouveau mot de passe
                new_password = st.text_input("Nouveau mot de passe (laisser vide pour générer automatiquement):", type="password", key="new_password")
                
                if st.button("🔒 Modifier le mot de passe", type="primary"):
                    if not new_password:
                        # Récupérer prénom et nom pour générer le mot de passe
                        firstname = user_data['Prénom']
                        lastname = user_data['Nom'] 
                        new_password = generate_password(firstname, lastname)
                        st.info(f"Mot de passe généré automatiquement: **{new_password}**")
                    
                    # Commande pour modifier le mot de passe
                    username = user_data['Login']
                    cmd_password = f"sudo -S samba-tool user setpassword {username} --newpassword={new_password}"
                    
                    try:
                        with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                            if client:
                                output, error = execute_ssh_command(client, cmd_password, config.SAMBA_PWD)
                                if error:
                                    st.error(f"❌ Erreur lors de la modification: {error}")
                                else:
                                    # Mettre à jour le fichier Excel
                                    df_users.loc[selected_index, 'Mot de passe'] = new_password
                                    df_users.loc[selected_index, 'Dernière modification'] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
                                    
                                    # Sauvegarder dans Excel
                                    df_users.to_excel(PASSWORD_EXCEL_FILE, index=False)
                                    
                                    st.success(f"✅ Mot de passe modifié pour {user_data['Prénom']} {user_data['Nom']}")
                                    st.success(f"**Nouveau mot de passe:** {new_password}")
                                    
                                    with st.expander("Logs de modification", expanded=False):
                                        st.code(output)
                            else:
                                st.error("❌ Impossible de se connecter au serveur Samba")
                    except Exception as e:
                        st.error(f"❌ Erreur : {str(e)}")
        elif not df_users.empty:
            st.error(f"Structure du fichier Excel incorrecte. Colonnes trouvées: {df_users.columns.tolist()}")
            st.info("Colonnes attendues: Prénom, Nom, Login, Mot de passe, Classe, Groupe")
        else:
            st.warning("Aucun utilisateur trouvé dans la base de données")

    # ========================= ONGLET 5: SUPPRESSION =========================
    with tab5:
        st.header("Suppression d'utilisateurs")
        
        # Sous-onglets pour différents types de suppression
        sub_tab1, sub_tab2 = st.tabs(["👤 Supprimer un utilisateur", "🧹 RAZ complète (Groupe Eleves)"])
        
        with sub_tab1:
            st.subheader("Supprimer un utilisateur spécifique")
            
            df_users = get_existing_users()
            
            if not df_users.empty and all(col in df_users.columns for col in ['Prénom', 'Nom', 'Login']):
                col1, col2 = st.columns(2)
                
                with col1:
                    # Sélection de l'utilisateur à supprimer
                    user_options = [f"{row['Prénom']} {row['Nom']} ({row['Login']})" for _, row in df_users.iterrows()]
                    selected_user_delete = st.selectbox("Sélectionner un utilisateur à supprimer:", user_options, key="delete_user_select")
                    
                    if selected_user_delete:
                        # Extraire les infos de l'utilisateur sélectionné
                        selected_index_delete = user_options.index(selected_user_delete)
                        user_data_delete = df_users.iloc[selected_index_delete]
                        
                        st.warning(f"**Utilisateur à supprimer:** {user_data_delete['Prénom']} {user_data_delete['Nom']}")
                        st.warning(f"**Login:** {user_data_delete['Login']}")
                        
                with col2:
                    st.error("⚠️ **ATTENTION:** Cette action est irréversible!")
                    
                    confirm_delete = st.checkbox("Je confirme vouloir supprimer cet utilisateur", key="confirm_single_delete")
                    
                    if confirm_delete and st.button("🗑️ Supprimer l'utilisateur", type="secondary"):
                        username_delete = user_data_delete['Login']
                        cmd_delete = f"sudo -S samba-tool user delete {username_delete}"
                        
                        try:
                            with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                                if client:
                                    output, error = execute_ssh_command(client, cmd_delete, config.SAMBA_PWD)
                                    if error:
                                        st.error(f"❌ Erreur lors de la suppression: {error}")
                                    else:
                                        # Supprimer de l'Excel
                                        df_users_updated = df_users.drop(selected_index_delete)
                                        df_users_updated.to_excel(PASSWORD_EXCEL_FILE, index=False)
                                        
                                        st.success(f"✅ Utilisateur {user_data_delete['Prénom']} {user_data_delete['Nom']} supprimé")
                                        
                                        with st.expander("Logs de suppression", expanded=False):
                                            st.code(output)
                                        
                                        st.rerun()
                                else:
                                    st.error("❌ Impossible de se connecter au serveur Samba")
                        except Exception as e:
                            st.error(f"❌ Erreur : {str(e)}")
            elif not df_users.empty:
                st.error(f"Structure du fichier Excel incorrecte. Colonnes trouvées: {df_users.columns.tolist()}")
                st.info("Colonnes attendues: Prénom, Nom, Login, Mot de passe, Classe, Groupe")
            else:
                st.warning("Aucun utilisateur trouvé dans la base de données")
        
        with sub_tab2:
            st.subheader("RAZ complète - Suppression du groupe Eleves")
            
            st.info("🔄 Récupération des membres réels du groupe 'Eleves' depuis Samba...")
            
            # Récupérer la liste réelle des membres du groupe Eleves depuis Samba
            eleves_members = get_samba_group_members("Eleves")
            eleves_count = len(eleves_members)
            
            st.error("🚨 **DANGER - REMISE À ZÉRO COMPLÈTE**")
            st.warning(f"Cette action va supprimer **{eleves_count} utilisateurs RÉELS** du groupe 'Eleves' dans Samba")
            
            if eleves_count > 0:
                with st.expander("Aperçu des utilisateurs à supprimer (depuis Samba)", expanded=False):
                    st.markdown("**Utilisateurs réellement présents dans le groupe 'Eleves' :**")
                    
                    # Créer un DataFrame avec les vrais utilisateurs
                    eleves_data = []
                    for username in eleves_members:
                        eleves_data.append({
                            "Login": username,
                            "Statut": "✅ Existe dans Samba"
                        })
                    
                    df_real_eleves = pd.DataFrame(eleves_data)
                    st.dataframe(df_real_eleves, hide_index=True, use_container_width=True)
                    
                    st.info("💡 Cette liste provient directement de Samba, pas du fichier Excel (plus fiable)")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.error("⚠️ **ATTENTION:** Cette action supprimera TOUS les comptes élèves!")
                    st.info("Utilisez cette fonction pour préparer la nouvelle année scolaire")
                    
                with col2:
                    confirm_raz = st.checkbox("Je confirme vouloir supprimer TOUS les comptes élèves", key="confirm_raz")
                    confirm_raz_text = st.text_input("Tapez 'SUPPRIMER TOUS' pour confirmer:", key="confirm_raz_text")
                    
                    if confirm_raz and confirm_raz_text == "SUPPRIMER TOUS":
                        if st.button("🧹 SUPPRIMER TOUS LES ÉLÈVES", type="secondary"):
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            results_raz = []
                            
                            # Utiliser la liste des vrais membres du groupe Eleves
                            for idx, username in enumerate(eleves_members):
                                status_text.text(f"Suppression de {username}... ({idx+1}/{eleves_count})")
                                
                                cmd_delete = f"sudo -S samba-tool user delete {username}"
                                
                                try:
                                    with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                                        if client:
                                            output, error = execute_ssh_command(client, cmd_delete, config.SAMBA_PWD)
                                            if error:
                                                results_raz.append(f"❌ {username}: {error}")
                                            else:
                                                results_raz.append(f"✅ {username} supprimé de Samba")
                                        else:
                                            results_raz.append(f"❌ {username}: Connexion impossible")
                                except Exception as e:
                                    results_raz.append(f"❌ {username}: {str(e)}")
                                
                                progress_bar.progress((idx + 1) / eleves_count)
                            
                            # Nettoyer le fichier Excel seulement pour les utilisateurs réellement supprimés
                            successfully_deleted = [r.split(": ")[0].replace("✅ ", "") for r in results_raz if "✅" in r]
                            
                            if successfully_deleted:
                                try:
                                    df_users = get_existing_users()  # Recharger le fichier Excel
                                    if not df_users.empty and 'Login' in df_users.columns:
                                        # Supprimer seulement les utilisateurs réellement supprimés de Samba
                                        df_remaining = df_users[~df_users['Login'].isin(successfully_deleted)]
                                        df_remaining.to_excel(PASSWORD_EXCEL_FILE, index=False)
                                        st.info(f"📝 {len(successfully_deleted)} utilisateurs supprimés du fichier Excel")
                                    else:
                                        st.warning("⚠️ Impossible de nettoyer le fichier Excel (structure invalide)")
                                except Exception as e:
                                    st.warning(f"⚠️ Erreur lors du nettoyage Excel : {e}")
                            
                            status_text.text("RAZ terminée!")
                            
                            st.subheader("Résultats de la suppression:")
                            success_count = sum(1 for r in results_raz if "✅" in r)
                            error_count = sum(1 for r in results_raz if "❌" in r)
                            
                            col_res1, col_res2 = st.columns(2)
                            with col_res1:
                                st.metric("✅ Supprimés avec succès", success_count)
                            with col_res2:
                                st.metric("❌ Erreurs", error_count)
                            
                            # Détail des résultats
                            with st.expander("Détails des opérations", expanded=False):
                                for result in results_raz:
                                    if "✅" in result:
                                        st.success(result)
                                    else:
                                        st.error(result)
                            
                            if success_count > 0:
                                st.success(f"✅ RAZ terminée - {success_count}/{eleves_count} comptes supprimés avec succès")
                            else:
                                st.error("❌ Aucun compte n'a pu être supprimé")
            else:
                st.info("Aucun utilisateur dans le groupe 'Eleves' à supprimer")

    # ========================= ONGLET 6: GROUPES =========================
    with tab6:
        st.header("Gestion des groupes")
        
        # Section Gestion du groupe WIFI
        st.subheader("📶 Gestion du groupe WIFI")
        st.markdown("**Ajoutez automatiquement des élèves au groupe WIFI selon leur classe/description**")
        
        # Interface pour sélectionner les descriptions
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Descriptions à rechercher :**")
            descriptions_input = st.text_area(
                "Entrez les descriptions (une par ligne)",
                value="TBAC\nBTS\nTFP",
                height=100,
                help="Entrez les mots-clés à rechercher dans les classes (ex: TBAC, BTS, TFP, etc.)"
            )
            
            # Nom du groupe cible
            target_group = st.text_input(
                "Nom du groupe cible",
                value="WIFI",
                help="Le groupe Samba où ajouter les élèves"
            )
        
        with col2:
            # Aperçu des utilisateurs qui correspondent
            if descriptions_input:
                descriptions_list = [desc.strip() for desc in descriptions_input.split('\n') if desc.strip()]
                
                if descriptions_list:
                    st.markdown("**Aperçu des utilisateurs correspondants :**")
                    
                    # Récupérer les utilisateurs pour l'aperçu
                    try:
                        df_users = get_existing_users()
                        if not df_users.empty and 'Classe' in df_users.columns:
                            matching_preview = []
                            for _, user in df_users.iterrows():
                                user_classe = str(user['Classe']).strip().upper()
                                for desc in descriptions_list:
                                    if desc.upper() in user_classe:
                                        matching_preview.append({
                                            "Login": user['Login'],
                                            "Nom": user.get('Nom', 'N/A'),
                                            "Prénom": user.get('Prénom', 'N/A'),
                                            "Classe": user['Classe']
                                        })
                                        break
                            
                            if matching_preview:
                                df_preview = pd.DataFrame(matching_preview)
                                st.dataframe(df_preview, hide_index=True, use_container_width=True)
                                st.info(f"🔍 {len(matching_preview)} utilisateurs trouvés")
                            else:
                                st.warning("Aucun utilisateur trouvé avec ces descriptions")
                        else:
                            st.warning("Aucun utilisateur dans la base de données")
                    except Exception as e:
                        st.error(f"Erreur lors de l'aperçu : {e}")
        
        # Bouton d'exécution
        if st.button("📶 Ajouter au groupe WIFI", type="primary"):
            if descriptions_input and target_group:
                descriptions_list = [desc.strip() for desc in descriptions_input.split('\n') if desc.strip()]
                
                if descriptions_list:
                    with st.spinner(f"Ajout des utilisateurs au groupe {target_group}..."):
                        results, added_count = add_students_to_wifi_group_by_description(descriptions_list, target_group)
                    
                    # Affichage des résultats
                    if results:
                        col_res1, col_res2 = st.columns(2)
                        with col_res1:
                            st.metric("✅ Ajoutés avec succès", added_count)
                        with col_res2:
                            total_processed = len(results)
                            st.metric("📊 Total traité", total_processed)
                        
                        # Détail des résultats
                        st.markdown("**Détails des opérations :**")
                        for result in results:
                            if "✅" in result:
                                st.success(result)
                            elif "ℹ️" in result:
                                st.info(result)
                            else:
                                st.error(result)
                        
                        if added_count > 0:
                            st.success(f"🎉 {added_count} utilisateur(s) ajouté(s) au groupe {target_group} avec succès!")
                        else:
                            st.info("Aucun nouvel utilisateur ajouté (tous déjà membres ou erreurs)")
                    else:
                        st.warning("Aucun résultat retourné")
                else:
                    st.error("Veuillez saisir au moins une description")
            else:
                st.error("Veuillez remplir tous les champs requis")

    # ========================= ONGLET 7: SYNC IMFR/SAMBA =========================
    with tab7:
        st.header("Synchronisation IMFR / SAMBA")
        st.markdown("**Comparez et synchronisez les listes d'élèves entre IMFR et SAMBA**")

        # Chargement automatique des élèves IMFR au démarrage
        if 'imfr_data' not in st.session_state:
            try:
                import pymysql
                conn = pymysql.connect(**config.MYSQL_CONFIG)
                query = """
                SELECT
                    nom as 'Nom',
                    prenom as 'Prénom',
                    classe as 'Classe'
                FROM eleves_imfr
                ORDER BY classe, nom, prenom
                """
                df_imfr = pd.read_sql(query, conn)
                conn.close()

                if not df_imfr.empty:
                    st.session_state.imfr_data = df_imfr
                    st.success(f"✅ {len(df_imfr)} élèves IMFR chargés automatiquement depuis MySQL")
            except:
                pass  # Silencieux si la table n'existe pas encore

        # Chargement automatique des utilisateurs SAMBA au démarrage
        if 'samba_data' not in st.session_state:
            try:
                import pymysql
                conn = pymysql.connect(**config.MYSQL_CONFIG)
                query = """
                SELECT
                    Nom as 'Nom',
                    Prénom as 'Prénom',
                    Classe as 'Classe',
                    Login as 'Login'
                FROM utilisateurs
                WHERE Login IS NOT NULL AND Login != ''
                ORDER BY Classe, Nom, Prénom
                """
                df_samba = pd.read_sql(query, conn)
                conn.close()

                if not df_samba.empty:
                    st.session_state.samba_data = df_samba
                    st.success(f"✅ {len(df_samba)} utilisateurs SAMBA chargés automatiquement depuis la table 'utilisateurs'")
            except Exception as e:
                pass  # Silencieux si la table n'existe pas encore

        # Section de récupération et actualisation
        st.subheader("🔄 Actualisation des données")

        col_refresh1, col_refresh2, col_refresh3 = st.columns(3)

        with col_refresh1:
            st.markdown("**💾 Base de données MySQL**")
            if st.button("🔄 Recharger IMFR depuis MySQL"):
                try:
                    import pymysql
                    conn = pymysql.connect(**config.MYSQL_CONFIG)
                    query = """
                    SELECT
                        nom as 'Nom',
                        prenom as 'Prénom',
                        classe as 'Classe'
                    FROM eleves_imfr
                    ORDER BY classe, nom, prenom
                    """
                    df_imfr = pd.read_sql(query, conn)
                    conn.close()

                    if not df_imfr.empty:
                        st.session_state.imfr_data = df_imfr
                        st.success(f"✅ {len(df_imfr)} élèves IMFR rechargés")
                    else:
                        st.warning("⚠️ Aucun élève dans la base de données MySQL")

                except pymysql.Error as e:
                    if "doesn't exist" in str(e):
                        st.error("❌ Table 'eleves_imfr' n'existe pas")
                        st.info("💡 Utilisez l'app 'Récupération Liste Élèves'")
                    else:
                        st.error(f"❌ Erreur MySQL: {e}")

        with col_refresh2:
            st.markdown("**🕷️ Scraping IMFR Direct**")
            if st.button("🔄 Lancer le scraping IMFR", help="Récupère directement depuis IMFR et sauvegarde dans MySQL"):
                with st.spinner("Scraping IMFR en cours (30-90s)..."):
                    try:
                        # Import du module de scraping
                        sys.path.insert(0, '/home/streamlit/apps')
                        from recuperation_eleves import login_to_site, navigate_to_section, go_to_excel_et_fusion

                        # Lancer le scraping
                        driver = login_to_site(config.IMFR_CONFIG)
                        navigate_to_section(driver)
                        time.sleep(2)
                        eleves_data_raw = go_to_excel_et_fusion(driver)
                        driver.quit()

                        if eleves_data_raw:
                            # Sauvegarder dans MySQL
                            import pymysql
                            conn = pymysql.connect(**config.MYSQL_CONFIG)
                            cursor = conn.cursor()

                            # Créer la table si nécessaire
                            create_table_query = """
                            CREATE TABLE IF NOT EXISTS eleves_imfr (
                                id INT AUTO_INCREMENT PRIMARY KEY,
                                nom VARCHAR(100) NOT NULL,
                                prenom VARCHAR(100) NOT NULL,
                                classe VARCHAR(50),
                                date_import TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                INDEX idx_nom_prenom (nom, prenom),
                                INDEX idx_classe (classe)
                            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                            """
                            cursor.execute(create_table_query)
                            cursor.execute("TRUNCATE TABLE eleves_imfr")

                            # Insérer les données
                            insert_query = "INSERT INTO eleves_imfr (nom, prenom, classe) VALUES (%s, %s, %s)"
                            for eleve in eleves_data_raw:
                                cursor.execute(insert_query, (eleve['nom'], eleve['prenom'], eleve['classe']))

                            conn.commit()
                            cursor.close()
                            conn.close()

                            # Recharger dans la session
                            df_imfr = pd.DataFrame({
                                'Nom': [e['nom'] for e in eleves_data_raw],
                                'Prénom': [e['prenom'] for e in eleves_data_raw],
                                'Classe': [e['classe'] for e in eleves_data_raw]
                            })
                            st.session_state.imfr_data = df_imfr

                            st.success(f"✅ {len(eleves_data_raw)} élèves récupérés depuis IMFR et sauvegardés dans MySQL")
                        else:
                            st.error("❌ Aucune donnée récupérée")

                    except Exception as e:
                        st.error(f"❌ Erreur scraping IMFR: {e}")

        with col_refresh3:
            st.markdown("**🔐 Mise à jour SAMBA**")
            if st.button("🔄 Mettre à jour depuis SAMBA", help="Met à jour la table utilisateurs depuis SAMBA via samba-tool user list"):
                with st.spinner("Récupération depuis SAMBA via SSH..."):
                    try:
                        # Se connecter au serveur SAMBA via SSH
                        with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                            if client:
                                # Exécuter samba-tool user list
                                cmd = "sudo -S samba-tool user list"
                                output, error = execute_ssh_command(client, cmd, config.SAMBA_PWD)

                                if output:
                                    # Parser les utilisateurs
                                    users = []
                                    ignored_users = config.get_ignored_users()

                                    # Liste de mots-clés à ignorer dans la sortie
                                    ignore_keywords = [
                                        'sudo', 'password', 'mot de passe', 'warning',
                                        'note:', 'ldb_wrap', 'service', 'flagged',
                                        'making it unavailable', 'path in service'
                                    ]

                                    for line in output.split('\n'):
                                        username = line.strip()

                                        # Ignorer les lignes vides ou trop courtes
                                        if not username or len(username) < 3:
                                            continue

                                        # Ignorer les lignes avec des mots-clés système
                                        if any(keyword.lower() in username.lower() for keyword in ignore_keywords):
                                            continue

                                        # Ignorer les lignes qui contiennent ':' (prompts système)
                                        if ':' in username and not '.' in username:
                                            continue

                                        # Ignorer les utilisateurs dans la liste d'exclusion
                                        if username.lower() not in [u.lower() for u in ignored_users]:
                                            users.append(username)

                                    st.info(f"📋 {len(users)} utilisateurs récupérés depuis SAMBA")

                                    # Pour chaque utilisateur, récupérer les détails
                                    users_details = []
                                    progress_bar = st.progress(0)

                                    for idx, username in enumerate(users):
                                        # Récupérer les détails de l'utilisateur
                                        cmd_details = f"sudo -S samba-tool user show {username}"
                                        details_output, details_error = execute_ssh_command(client, cmd_details, config.SAMBA_PWD)

                                        if details_output:
                                            # Parser les détails (nom, prénom, etc.)
                                            user_info = {
                                                'Login': username,
                                                'Nom': '',
                                                'Prénom': '',
                                                'Classe': '',
                                                'Groupe': ''
                                            }

                                            # Parser les attributs LDAP
                                            for line in details_output.split('\n'):
                                                if ':' in line:
                                                    key, value = line.split(':', 1)
                                                    key = key.strip()
                                                    value = value.strip()

                                                    if key == 'sn':  # Nom
                                                        user_info['Nom'] = value
                                                    elif key == 'givenName':  # Prénom
                                                        user_info['Prénom'] = value
                                                    elif key == 'description' or key == 'comment':
                                                        # Extraire classe si présente
                                                        if value and not any(kw in value.lower() for kw in ['warning', 'note', 'sudo']):
                                                            user_info['Classe'] = value
                                                    elif key == 'memberOf':
                                                        # Extraire le groupe principal (premier CN=)
                                                        if 'CN=' in value:
                                                            group = value.split('CN=')[1].split(',')[0]
                                                            if group and not any(kw in group.lower() for kw in ['users', 'domain']):
                                                                user_info['Groupe'] = group

                                            # Si sn et givenName ne sont pas trouvés, extraire depuis le login
                                            # Format attendu: prenom.nom
                                            if not user_info['Nom'] and not user_info['Prénom']:
                                                if '.' in username:
                                                    parts = username.split('.')
                                                    if len(parts) >= 2:
                                                        user_info['Prénom'] = parts[0].capitalize()
                                                        user_info['Nom'] = parts[1].upper()

                                            users_details.append(user_info)

                                        progress_bar.progress((idx + 1) / len(users))

                                    progress_bar.empty()

                                    if users_details:
                                        # Mettre à jour la table utilisateurs dans MySQL
                                        import pymysql
                                        conn = pymysql.connect(**config.MYSQL_CONFIG)
                                        cursor = conn.cursor()

                                        # Marquer tous les utilisateurs existants comme non-actifs
                                        update_query = """
                                        UPDATE utilisateurs
                                        SET Dernière_modification = NOW()
                                        WHERE Login IN (SELECT Login FROM (SELECT Login FROM utilisateurs) AS temp)
                                        """

                                        # Insérer ou mettre à jour les utilisateurs
                                        upsert_query = """
                                        INSERT INTO utilisateurs (Login, Nom, Prénom, Classe, Groupe, Mot_de_passe, Dernière_modification)
                                        VALUES (%s, %s, %s, %s, %s, '****', NOW())
                                        ON DUPLICATE KEY UPDATE
                                            Nom = VALUES(Nom),
                                            Prénom = VALUES(Prénom),
                                            Classe = VALUES(Classe),
                                            Groupe = VALUES(Groupe),
                                            Dernière_modification = NOW()
                                        """

                                        count_updated = 0
                                        for user in users_details:
                                            cursor.execute(upsert_query, (
                                                user['Login'],
                                                user['Nom'],
                                                user['Prénom'],
                                                user['Classe'],
                                                user['Groupe']
                                            ))
                                            count_updated += 1

                                        conn.commit()
                                        cursor.close()
                                        conn.close()

                                        # Recharger les données pour l'affichage
                                        conn = pymysql.connect(**config.MYSQL_CONFIG)
                                        query = """
                                        SELECT
                                            Login,
                                            Nom,
                                            Prénom,
                                            Classe,
                                            Groupe
                                        FROM utilisateurs
                                        ORDER BY Nom, Prénom
                                        """
                                        df_samba = pd.read_sql(query, conn)
                                        conn.close()

                                        st.session_state.samba_data = df_samba
                                        st.success(f"✅ {count_updated} utilisateurs synchronisés dans la table 'utilisateurs'")

                                    else:
                                        st.warning("⚠️ Aucun détail utilisateur récupéré")

                                else:
                                    st.error(f"❌ Erreur lors de l'exécution de samba-tool: {error}")

                            else:
                                st.error("❌ Impossible de se connecter au serveur SAMBA")

                    except Exception as e:
                        st.error(f"❌ Erreur récupération SAMBA: {e}")
                        import traceback
                        st.error(f"Détails: {traceback.format_exc()}")

        st.markdown("---")

        # Afficher les statistiques des données chargées
        st.subheader("📊 Données chargées")

        col_stat1, col_stat2 = st.columns(2)

        with col_stat1:
            if 'imfr_data' in st.session_state and not st.session_state.imfr_data.empty:
                df_imfr = st.session_state.imfr_data
                st.markdown("**📘 Élèves IMFR**")
                st.metric("Total élèves IMFR", len(df_imfr))
                nb_classes = df_imfr['Classe'].nunique() if 'Classe' in df_imfr.columns else 0
                st.metric("Nombre de classes", nb_classes)

                with st.expander("Aperçu des élèves IMFR", expanded=False):
                    st.dataframe(df_imfr.head(20), hide_index=True)
            else:
                st.info("ℹ️ Aucune donnée IMFR chargée")

        with col_stat2:
            if 'samba_data' in st.session_state and not st.session_state.samba_data.empty:
                df_samba = st.session_state.samba_data
                st.markdown("**🔐 Utilisateurs SAMBA**")
                st.metric("Total utilisateurs SAMBA", len(df_samba))

                with st.expander("Aperçu des utilisateurs SAMBA", expanded=False):
                    st.dataframe(df_samba.head(20), hide_index=True)
            else:
                st.info("ℹ️ Aucune donnée SAMBA chargée")

        st.markdown("---")

        # Section de comparaison
        if 'imfr_data' in st.session_state and 'samba_data' in st.session_state:
            st.subheader("🔍 Comparaison IMFR vs SAMBA")

            df_imfr = st.session_state.imfr_data
            df_samba = st.session_state.samba_data

            if not df_samba.empty and not df_imfr.empty:
                # Comparaison des données
                col_comp1, col_comp2 = st.columns(2)
                
                with col_comp1:
                    st.markdown("**📊 Statistiques :**")
                    st.metric("Élèves dans IMFR", len(df_imfr))
                    st.metric("Utilisateurs dans SAMBA", len(df_samba))
                
                # Trouver les élèves manquants dans SAMBA
                missing_in_samba = []
                for _, imfr_student in df_imfr.iterrows():
                    imfr_nom = str(imfr_student['Nom']).strip().upper()
                    imfr_prenom = str(imfr_student['Prénom']).strip().title()
                    
                    # Chercher dans SAMBA
                    found = False
                    for _, samba_student in df_samba.iterrows():
                        samba_nom = str(samba_student.get('Nom', '')).strip().upper()
                        samba_prenom = str(samba_student.get('Prénom', '')).strip().title()
                        
                        if imfr_nom == samba_nom and imfr_prenom == samba_prenom:
                            found = True
                            break
                    
                    if not found:
                        missing_in_samba.append(imfr_student)
                
                # Trouver les élèves en plus dans SAMBA
                extra_in_samba = []
                for _, samba_student in df_samba.iterrows():
                    samba_nom = str(samba_student.get('Nom', '')).strip().upper()
                    samba_prenom = str(samba_student.get('Prénom', '')).strip().title()
                    
                    # Chercher dans IMFR
                    found = False
                    for _, imfr_student in df_imfr.iterrows():
                        imfr_nom = str(imfr_student['Nom']).strip().upper()
                        imfr_prenom = str(imfr_student['Prénom']).strip().title()
                        
                        if samba_nom == imfr_nom and samba_prenom == imfr_prenom:
                            found = True
                            break
                    
                    if not found:
                        extra_in_samba.append(samba_student)
                
                with col_comp2:
                    st.metric("Manquants dans SAMBA", len(missing_in_samba))
                    st.metric("En plus dans SAMBA", len(extra_in_samba))
                
                st.markdown("---")
                
                # Section pour créer les élèves manquants
                if missing_in_samba:
                    st.subheader("➕ Élèves à créer dans SAMBA")

                    df_missing = pd.DataFrame(missing_in_samba)

                    # Section d'exclusion des élèves
                    with st.expander("🚫 Exclure des élèves de la création", expanded=False):
                        st.markdown("**Sélectionnez les élèves à NE PAS créer dans SAMBA :**")

                        # Initialiser la liste d'exclusion dans session_state si elle n'existe pas
                        if 'excluded_students' not in st.session_state:
                            st.session_state.excluded_students = []

                        # Créer une clé unique pour chaque élève (Nom + Prénom)
                        for idx, student in enumerate(missing_in_samba):
                            student_key = f"{student['Nom']}_{student['Prénom']}"
                            student_label = f"{student['Prénom']} {student['Nom']} ({student.get('Classe', 'N/A')})"

                            # Checkbox pour exclure cet élève
                            is_excluded = st.checkbox(
                                f"🚫 Exclure : {student_label}",
                                value=student_key in st.session_state.excluded_students,
                                key=f"exclude_{idx}_{student_key}"
                            )

                            # Mettre à jour la liste d'exclusion
                            if is_excluded and student_key not in st.session_state.excluded_students:
                                st.session_state.excluded_students.append(student_key)
                            elif not is_excluded and student_key in st.session_state.excluded_students:
                                st.session_state.excluded_students.remove(student_key)

                        # Boutons de gestion des exclusions
                        col_excl1, col_excl2, col_excl3 = st.columns(3)
                        with col_excl1:
                            if st.button("✅ Tout sélectionner"):
                                st.session_state.excluded_students = [f"{s['Nom']}_{s['Prénom']}" for s in missing_in_samba]
                                st.rerun()
                        with col_excl2:
                            if st.button("❌ Tout désélectionner"):
                                st.session_state.excluded_students = []
                                st.rerun()
                        with col_excl3:
                            nb_excluded = len(st.session_state.excluded_students)
                            st.metric("Élèves exclus", nb_excluded)

                    # Filtrer les élèves manquants en excluant ceux qui sont cochés
                    filtered_missing = [
                        s for s in missing_in_samba
                        if f"{s['Nom']}_{s['Prénom']}" not in st.session_state.excluded_students
                    ]

                    # Afficher le tableau des élèves à créer (après exclusion)
                    st.markdown(f"**Élèves à créer : {len(filtered_missing)} / {len(missing_in_samba)}**")
                    if filtered_missing:
                        df_filtered = pd.DataFrame(filtered_missing)
                        st.dataframe(df_filtered[['Nom', 'Prénom', 'Classe']], hide_index=True, use_container_width=True)
                    else:
                        st.info("ℹ️ Tous les élèves ont été exclus. Aucun compte ne sera créé.")

                    col_create1, col_create2 = st.columns(2)
                    
                    with col_create1:
                        groupe_cible = st.selectbox(
                            "Groupe SAMBA pour les nouveaux élèves",
                            ["Eleves", "WIFI", "Autre"],
                            index=0
                        )
                        
                        if groupe_cible == "Autre":
                            groupe_cible = st.text_input("Nom du groupe personnalisé", value="Eleves")
                    
                    with col_create2:
                        generate_passwords = st.checkbox("Générer des mots de passe automatiquement", value=True)
                        send_to_db = st.checkbox("Sauvegarder dans la base de données", value=True, help="Enregistre dans la base de données MySQL (table utilisateurs)")
                    
                    if st.button("🔨 Créer tous les élèves manquants", type="primary"):
                        if filtered_missing:
                            creation_results = []
                            progress_bar = st.progress(0)
                            status_text = st.empty()

                            for idx, student in enumerate(filtered_missing):
                                nom = str(student['Nom']).strip()
                                prenom = str(student['Prénom']).strip()
                                classe = str(student['Classe']).strip()
                                
                                status_text.text(f"Création de {prenom} {nom}... ({idx+1}/{len(filtered_missing)})")
                                
                                # Générer username et password
                                username = generate_username(prenom, nom)
                                if generate_passwords:
                                    password = generate_password()
                                else:
                                    password = f"{nom.lower()}{prenom.lower()[0]}{random.randint(10,99)}"
                                
                                # Créer l'utilisateur SAMBA (utiliser les fonctions existantes)
                                try:
                                    # Normaliser la classe
                                    classe_normalized = normalize_class_name(classe) if classe else ""
                                    
                                    # Commandes de création SAMBA
                                    user_commands = []
                                    group_commands = []
                                    
                                    # Commande de création d'utilisateur
                                    user_commands.append(
                                        f"sudo -S samba-tool user create {username} '{password}' "
                                        f"--given-name='{prenom}' --surname='{nom}' "
                                        f"--description='{classe_normalized}'"
                                    )
                                    
                                    # Commande d'ajout au groupe
                                    group_commands.append(f"sudo -S samba-tool group addmembers {groupe_cible} {username}")
                                    
                                    # Exécuter les commandes
                                    with ssh_connection(config.SAMBA_SERVER, config.SAMBA_USER, config.SAMBA_PWD) as client:
                                        if client:
                                            # Créer l'utilisateur
                                            for cmd in user_commands:
                                                output, error = execute_ssh_command(client, cmd, config.SAMBA_PWD)
                                                if error:
                                                    raise Exception(f"Erreur création utilisateur: {error}")
                                            
                                            # Ajouter au groupe
                                            for cmd in group_commands:
                                                output, error = execute_ssh_command(client, cmd, config.SAMBA_PWD)
                                                if error and "is already a member" not in error.lower():
                                                    st.warning(f"Avertissement groupe pour {username}: {error}")
                                            
                                            result = f"✅ {prenom} {nom} ({classe_normalized}) : Utilisateur créé avec succès"
                                            creation_results.append(result)

                                            # Sauvegarder dans la base de données si demandé
                                            if send_to_db:
                                                save_user_to_db(username, prenom, nom, password, classe_normalized, groupe_cible)
                                        else:
                                            raise Exception("Impossible de se connecter au serveur SAMBA")
                                    
                                except Exception as e:
                                    result = f"❌ {prenom} {nom} : {str(e)}"
                                    creation_results.append(result)
                                
                                progress_bar.progress((idx + 1) / len(filtered_missing))
                            
                            status_text.text("Création terminée!")
                            
                            # Afficher les résultats
                            st.subheader("Résultats de la création :")
                            success_count = sum(1 for r in creation_results if "✅" in r)
                            error_count = sum(1 for r in creation_results if "❌" in r)
                            
                            col_res1, col_res2 = st.columns(2)
                            with col_res1:
                                st.metric("✅ Créés avec succès", success_count)
                            with col_res2:
                                st.metric("❌ Erreurs", error_count)
                            
                            # Détails
                            with st.expander("Détails des créations", expanded=True):
                                for result in creation_results:
                                    if "✅" in result:
                                        st.success(result)
                                    else:
                                        st.error(result)
                            
                            if success_count > 0:
                                st.success(f"🎉 {success_count} élève(s) créé(s) avec succès dans SAMBA!")
                                # Rafraîchir les données
                                if st.button("🔄 Actualiser la comparaison"):
                                    st.rerun()
                        else:
                            st.warning("⚠️ Aucun élève à créer. Tous les élèves ont été exclus ou sont déjà présents dans SAMBA.")
                else:
                    st.success("✅ Tous les élèves d'IMFR sont déjà présents dans SAMBA!")
                
                # Section pour les élèves en plus dans SAMBA
                if extra_in_samba:
                    st.markdown("---")
                    st.subheader("⚠️ Élèves présents uniquement dans SAMBA")
                    st.markdown("*Ces élèves sont dans SAMBA mais pas dans IMFR*")
                    
                    df_extra = pd.DataFrame(extra_in_samba)
                    required_columns = ['Nom', 'Prénom', 'Login']
                    display_columns = [col for col in required_columns if col in df_extra.columns]
                    
                    if 'Classe' in df_extra.columns:
                        display_columns.append('Classe')
                    
                    st.dataframe(df_extra[display_columns], hide_index=True, use_container_width=True)
                    
                    st.info("💡 Ces comptes peuvent être des anciens élèves ou des comptes créés manuellement.")
            else:
                st.warning("Aucun utilisateur trouvé dans SAMBA")
        else:
            st.info("📋 Veuillez d'abord charger les données IMFR pour effectuer la comparaison")
    with tab8:
        st.header("Outils et diagnostics")
        
        # Section Correspondance des classes
        st.subheader("📋 Correspondance des classes")
        with st.expander("Tableau de correspondance complet", expanded=False):
            st.markdown("**Correspondances automatiques des noms de classes :**")
            
            # Créer un DataFrame pour un affichage plus propre
            mapping_data = []
            for original, normalized in CLASS_MAPPING.items():
                mapping_data.append({"Nom original": original, "Abréviation": normalized})
            
            df_mapping = pd.DataFrame(mapping_data)
            st.dataframe(df_mapping, hide_index=True, use_container_width=True)
            
            st.info(config.HELP_MESSAGES['class_mapping'])
            
            # Bouton d'export du mapping
            csv_mapping = df_mapping.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📥 Télécharger le tableau de correspondance CSV",
                data=csv_mapping,
                file_name="correspondance_classes.csv",
                mime="text/csv"
            )
        
        st.markdown("---")
        
        # Section Licences Microsoft 365 disponibles
        st.subheader("🔑 Licences Microsoft 365 disponibles")
        with st.expander("Consulter les licences dans Entra ID", expanded=False):
            with st.spinner("📡 Récupération des licences depuis Microsoft Graph..."):
                available_licenses = get_available_licenses()
            
            if available_licenses:
                st.success(f"✅ {len(available_licenses)} types de licences trouvés dans Entra ID")
                
                # Créer un DataFrame pour l'affichage
                licenses_data = []
                for sku_id, license_info in available_licenses.items():
                    licenses_data.append({
                        "Nom de la licence": license_info['displayName'],
                        "Code SKU": license_info['partNumber'],
                        "Unités activées": license_info['enabledUnits'],
                        "Unités consommées": license_info['consumedUnits'],
                        "Unités disponibles": license_info['availableUnits']
                    })
                
                df_licenses = pd.DataFrame(licenses_data)
                st.dataframe(df_licenses, hide_index=True, use_container_width=True)
                
                # Bouton d'export des licences
                csv_licenses = df_licenses.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📥 Télécharger la liste des licences CSV",
                    data=csv_licenses,
                    file_name="licences_microsoft365.csv",
                    mime="text/csv"
                )
            else:
                st.warning("⚠️ Impossible de récupérer les licences depuis Entra ID")
                st.info("Vérifiez les permissions Microsoft Graph : Organization.Read.All")
        
        st.markdown("---")
        
        # Section Groupes Entra ID réels
        st.subheader("👥 Groupes Entra ID (Azure AD)")
        with st.expander("Consulter les groupes dans Entra ID", expanded=False):
            with st.spinner("📡 Récupération des groupes depuis Entra ID..."):
                entra_groups = get_entra_groups()
            
            if entra_groups:
                st.success(f"✅ {len(entra_groups)} groupes trouvés dans Entra ID")
                
                # Filtres pour les groupes
                col_filter1, col_filter2, col_filter3 = st.columns(3)
                
                with col_filter1:
                    filter_type = st.selectbox(
                        "Filtrer par type:",
                        ["Tous", "Microsoft 365", "Sécurité", "Distribution", "Sécurité avec messagerie", "Autre"],
                        key="group_type_filter"
                    )
                
                with col_filter2:
                    search_term = st.text_input("Rechercher par nom:", key="group_search")
                
                with col_filter3:
                    show_all = st.checkbox("Afficher tous les détails", key="show_group_details")
                
                # Appliquer les filtres
                filtered_groups = []
                for group_id, group_info in entra_groups.items():
                    # Filtre par type
                    if filter_type != "Tous" and group_info['groupType'] != filter_type:
                        continue
                    
                    # Filtre par recherche
                    if search_term and search_term.lower() not in group_info['displayName'].lower():
                        continue
                    
                    filtered_groups.append({
                        "Nom du groupe": group_info['displayName'],
                        "Type": group_info['groupType'],
                        "Description": group_info['description'] if group_info['description'] else "Aucune",
                        "ID Azure": group_id if show_all else group_id[:8] + "...",
                        "Sécurité": "✅" if group_info['securityEnabled'] else "❌",
                        "Messagerie": "✅" if group_info['mailEnabled'] else "❌"
                    })
                
                if filtered_groups:
                    df_entra_groups = pd.DataFrame(filtered_groups)
                    st.dataframe(df_entra_groups, hide_index=True, use_container_width=True)
                    
                    # Bouton d'export des groupes
                    csv_groups = df_entra_groups.to_csv(index=False, encoding='utf-8-sig')
                    st.download_button(
                        label="📥 Télécharger la liste des groupes CSV",
                        data=csv_groups,
                        file_name="groupes_entra_id.csv",
                        mime="text/csv"
                    )
                    
                    st.info(f"📊 {len(filtered_groups)} groupes affichés sur {len(entra_groups)} total")
                else:
                    st.warning("Aucun groupe ne correspond aux filtres sélectionnés")
                    
                # Recherche de groupes spécifiques configurés
                st.markdown("---")
                st.markdown("**🔍 Vérification des groupes configurés :**")
                
                configured_groups = {
                    config.LICENSE_GROUP_STUDENTS: config.LICENSE_GROUP_STUDENTS_ID,
                    config.LICENSE_GROUP_TEACHERS: None,
                    config.LICENSE_GROUP_OFFICE: None
                }
                
                for group_name, expected_id in configured_groups.items():
                    found_groups = [g for g in entra_groups.values() if g['displayName'] == group_name]
                    
                    if found_groups:
                        group = found_groups[0]
                        status = "✅ Trouvé"
                        id_match = group['id'] == expected_id if expected_id else "Non défini"
                        st.success(f"**{group_name}**: {status} - ID: `{group['id']}` - Type: {group['groupType']}")
                    else:
                        st.warning(f"**{group_name}**: ⚠️ Non trouvé dans Entra ID")
            else:
                st.warning("⚠️ Impossible de récupérer les groupes depuis Entra ID")
                st.info("Vérifiez les permissions Microsoft Graph : Group.Read.All ou Directory.Read.All")
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Diagnostics système")
            if st.button("🔍 Diagnostiquer l'environnement"):
                diagnose_environment()
        
        with col2:
            st.subheader("Tests de connexion")
            if st.button("🧪 Tester la commande de sync"):
                test_sync_command()
        
        st.markdown("---")
        st.info("💡 **Astuce:** Utilisez l'onglet 'Mots de Passe' dans le menu principal pour une vue d'ensemble de tous les comptes utilisateurs.")