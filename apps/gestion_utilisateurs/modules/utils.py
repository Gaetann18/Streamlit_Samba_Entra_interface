"""
Module des fonctions utilitaires pour la gestion des utilisateurs
"""
import streamlit as st
import subprocess
import random
from contextlib import contextmanager
import paramiko
import sys
import os
import logging
import time
import pandas as pd
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill
import uuid
import requests
from msal import ConfidentialClientApplication

# Ajouter le chemin parent pour importer config
sys.path.append('/home/streamlit')
import config

# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Constantes depuis config.py
PASSWORD_EXCEL_FILE = config.PASSWORD_EXCEL_FILE
CLASS_MAPPING = config.CLASS_MAPPING


def generate_username(first_name, last_name):
    """Génère un nom d'utilisateur au format prenom.nom"""
    clean_first_name = first_name.strip().lower()
    clean_last_name = last_name.strip().lower()
    
    # Supprimer les accents et caractères spéciaux
    import unicodedata
    clean_first_name = unicodedata.normalize('NFD', clean_first_name).encode('ascii', 'ignore').decode('ascii')
    clean_last_name = unicodedata.normalize('NFD', clean_last_name).encode('ascii', 'ignore').decode('ascii')
    
    # Supprimer les espaces et caractères spéciaux
    clean_first_name = ''.join(c for c in clean_first_name if c.isalnum())
    clean_last_name = ''.join(c for c in clean_last_name if c.isalnum())
    
    # Format prenom.nom
    base_username = f"{clean_first_name}.{clean_last_name}"
    
    # Ajouter un suffixe numérique si nécessaire
    username = base_username
    counter = 1
    
    # Vérifier l'unicité
    existing_users = get_existing_users()
    if not existing_users.empty and 'Login' in existing_users.columns:
        existing_logins = existing_users['Login'].str.lower().tolist()
        
        while username.lower() in existing_logins:
            username = f"{clean_first_name}.{clean_last_name}{counter}"
            counter += 1
    
    return username


def generate_password(length=8):
    """Génère un mot de passe aléatoire"""
    characters = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    return ''.join(random.choice(characters) for _ in range(length))


def normalize_class_name(original_class):
    """Normalise le nom de classe selon le mapping défini"""
    if not original_class:
        return ""
    
    cleaned_class = " ".join(str(original_class).split())
    
    # Chercher dans le mapping
    if cleaned_class in CLASS_MAPPING:
        return CLASS_MAPPING[cleaned_class]
    
    # Traitement spécial pour les classes TFP
    if "TFP" in cleaned_class.upper():
        tfp_clean = cleaned_class.upper().replace("  ", " ").strip()
        
        if "ADMIN" in tfp_clean or "Admin" in cleaned_class:
            return "TFP ADMIN"
        elif "RAV" in tfp_clean or "Rav" in cleaned_class:
            return "TFP RAV"
        else:
            return "TFP"
    
    # Normalisation standard pour les autres classes
    normalized = cleaned_class.replace(" ", "").replace("è", "e").replace("nd", "").upper()
    logger.warning(f"Classe non mappée: '{original_class}' -> utilisation par défaut: '{normalized}'")
    
    return normalized


def get_existing_users():
    """Récupère les utilisateurs existants depuis le fichier Excel"""
    try:
        if os.path.exists(PASSWORD_EXCEL_FILE):
            df = pd.read_excel(PASSWORD_EXCEL_FILE)
            return df
        else:
            # Créer un DataFrame vide avec les colonnes attendues
            return pd.DataFrame(columns=['Prénom', 'Nom', 'Login', 'Mot de passe', 'Classe', 'Groupe'])
    except Exception as e:
        logger.error(f"Erreur lors de la lecture du fichier Excel: {e}")
        return pd.DataFrame(columns=['Prénom', 'Nom', 'Login', 'Mot de passe', 'Classe', 'Groupe'])


def save_user_to_excel(username, firstname, lastname, password, classe="", groupe="Eleves"):
    """Sauvegarde un utilisateur dans le fichier Excel"""
    try:
        # Normaliser la classe avant sauvegarde
        normalized_classe = normalize_class_name(classe) if classe else ""
        
        # Lire le fichier existant ou créer un nouveau
        if os.path.exists(PASSWORD_EXCEL_FILE):
            wb = openpyxl.load_workbook(PASSWORD_EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # Créer l'en-tête
            ws['A1'] = 'Prénom'
            ws['B1'] = 'Nom'
            ws['C1'] = 'Login'
            ws['D1'] = 'Mot de passe'
            ws['E1'] = 'Classe'
            ws['F1'] = 'Groupe'
        
        # Trouver la prochaine ligne disponible
        next_row = ws.max_row + 1
        
        # Ajouter les données
        ws[f'A{next_row}'] = firstname
        ws[f'B{next_row}'] = lastname
        ws[f'C{next_row}'] = username
        ws[f'D{next_row}'] = password
        ws[f'E{next_row}'] = normalized_classe
        ws[f'F{next_row}'] = groupe
        
        # Sauvegarder
        wb.save(PASSWORD_EXCEL_FILE)
        logger.info(f"Utilisateur {username} sauvegardé dans Excel")
        
    except Exception as e:
        logger.error(f"Erreur lors de la sauvegarde Excel: {e}")
        raise


@contextmanager
def ssh_connection(server, username, password):
    """Context manager pour les connexions SSH"""
    client = None
    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(server, username=username, password=password, timeout=30)
        yield client
    except Exception as e:
        logger.error(f"Erreur de connexion SSH: {e}")
        yield None
    finally:
        if client:
            try:
                client.close()
            except Exception as close_error:
                logger.error(f"Erreur lors de la fermeture SSH: {close_error}")


def execute_ssh_command(client, command, sudo_password=None, timeout=30):
    """Exécute une commande SSH avec gestion du sudo"""
    if client is None:
        return "", "Client SSH non connecté"
        
    try:
        if sudo_password and 'sudo -S' in command:
            stdin, stdout, stderr = client.exec_command(command, timeout=timeout)
            stdin.write(f"{sudo_password}\n")
            stdin.flush()
            stdin.close()  # Fermer stdin après écriture
        else:
            stdin, stdout, stderr = client.exec_command(command, timeout=timeout)
        
        # Attendre que la commande se termine avec timeout
        exit_status = stdout.channel.recv_exit_status()
        
        # Lire les sorties
        try:
            output = stdout.read().decode('utf-8')
            error = stderr.read().decode('utf-8')
        except UnicodeDecodeError:
            # Fallback en cas de problème d'encodage
            output = stdout.read().decode('utf-8', errors='ignore')
            error = stderr.read().decode('utf-8', errors='ignore')
        
        # Nettoyer les outputs
        output = output.strip()
        error = error.strip()
        
        return output, error
        
    except Exception as e:
        logger.error(f"Erreur lors de l'exécution de la commande '{command[:50]}...': {e}")
        return "", f"Erreur d'exécution: {str(e)}"


def update_users_created_session(new_users_data):
    """Met à jour la session avec les nouveaux utilisateurs créés"""
    if 'users_created' not in st.session_state:
        st.session_state.users_created = []
    
    st.session_state.users_created.extend(new_users_data)


def clear_users_created_session():
    """Vide la session des utilisateurs créés"""
    if 'users_created' in st.session_state:
        st.session_state.users_created = []